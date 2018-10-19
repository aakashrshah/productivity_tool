import os
import sys
import glob
import datetime
from shutil import copyfile
from datetime import date

import time 
import getpass 

import pandas as pd
import numpy as np

from tkinter import *
from tkinter import ttk
import tkinter.messagebox as messagebox
import tkinter.filedialog as filedialog
import tkinter.scrolledtext as scrolledtext

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

import matplotlib
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.animation as animation
from matplotlib import style
from matplotlib.ticker import MaxNLocator
import matplotlib.gridspec as gridspec

from collections import OrderedDict


import crud as db
import reports as report

def resource_path(relative_path):
	""" Get absolute path to resource, works for dev and for PyInstaller """
	try:
		# PyInstaller creates a temp folder and stores path in _MEIPASS
		base_path = sys._MEIPASS
	except Exception:
		base_path = os.path.abspath(".")

	return os.path.join(base_path, relative_path)


'''
Variables: 

'''
username = getpass.getuser() 
# user = aduser.ADUser.from_cn(username)
# user = user.get_attribute("displayName")
print(username)

default_dir = 'C:/Users/' + username + '/Documents/'
dbq = 'T:/CITDR/CITSQ/QA/Operational/Status Reports/QA reviews/Access_DB/CITSQ_Python_Tool/DB/QA Project Review_Database_v04 - Update test.mdb'
db.setup_conn(default_dir,dbq)

copied_folder = 'T:/CITDR/CITSQ/QA/Operational/Status Reports/QA reviews/Access_DB/CITSQ_Python_Tool/SOURCE'
image_loader_name = resource_path("./images/loader1.gif")

sheets = ['Test Plan', 'Test Cases', 'RTM', 'Test Results', 'Test Summary Report']

valid_column = 'Review Date'
count_columns_sheets = ['Test Cases','Test Results']
default_date = datetime.datetime(2001, 1, 1)
today = datetime.date.today()
now_date = today.strftime('%m/%d/%Y')
nan_date = datetime.date.min.strftime('%m/%d/%Y')
proj_artifacts = []
findings = []
pa_error = ''
finding_error = ''
new_project = False

# Configuration: row and column number
header_end = 7
qa_data_start = 7
TestCase_Finding_id_column = 26

missed_files = 0
# Configuration: data validation
valid_dict = {'Test Plan':"='Data Validations'!$A$2:$A$20",'Test Cases':"='Data Validations'!$E$2:$E$21",'Test Results':"='Data Validations'!$C$14:$C$25",
				   'RTM':"='Data Validations'!$C$2:$C$12",'Test Summary Report':"='Data Validations'!$C$31:$C$33"}
# empty review data row
epy_st = 10
last_row = 0
nan = np.nan
nan_str = 'nan'

# Number of new data(rows) create
count_new = 0
count_dict = {}

style.use("ggplot")
output_folder = "T:/CITDR/CITSQ/QA/Operational/Status Reports/QA reviews/Access_DB/CITSQ_Python_Tool/REPORTS"
if not os.path.exists("T:/CITDR/CITSQ/QA/Operational/Status Reports/QA reviews/Access_DB/CITSQ_Python_Tool/REPORTS"):
	os.makedirs(output_folder)

output_personalize = "\\" + str(username) + "\\" + str(today.strftime('%m-%d-%Y')) + "\\"
current_output_folder = output_folder + "/" + str(username) + "/" + str(today.strftime('%m-%d-%Y'))

if not os.path.exists(current_output_folder):
	os.makedirs(current_output_folder)


'''

Variables : End
Report : Han  

'''

class Report:
	def __init__(self, projects, location=None):
		if(location == None):
			print('QA Reporting System Init')
			# self.projects = projects
			if self.get_project_location(projects) == -1:
					return
			else:
				self.sync_findings_all(projects)
		else:
			if(len(projects) == 1):
				print('QA Reporting System with External File')
				if self.set_project_location(projects,location) == -1:
					print("Permission Error")
					return
				else:
					print("Sync Findings?")
					self.sync_findings_all_location(projects,location)
			else:
				print("Error")            
	
	def set_project_location(self, projects,location):
		self.project_location = {}
		project_name = projects[0]
		self.project_location[project_name] = location
		external_file = self.project_location[project_name]
		self.new_project_location = {}
		try:
			file_name = os.path.basename(external_file)
			copied_file_path = os.path.join(copied_folder, file_name)
			print(external_file, copied_file_path)
			# copyfile(external_file, copied_file_path)
			self.new_project_location[project_name] = copied_file_path
			print('New path: ' + copied_file_path)
			print('************* Copying project source files: Done')
		except PermissionError:
			return -1

	# Get the location of project from Database
	def get_project_location(self, projects):
		self.project_location = {}

		for proj in projects:
			self.project_location[proj] = db.query_projects_location(proj)

			# print(proj + ': ' + self.project_location[proj])

			if self.project_location[proj] == '':
				print(proj + ': No file path')
				sys.exit()
		try:
			self.copy_source()
		except PermissionError:
			return -1


	# Copy the project original files to the new folder
	def copy_source(self):
		global missed_files
		self.new_project_location = {}

		print('************* Copying project source files...')
		for proj in self.project_location:
			path = self.project_location[proj]

			print(proj)
			print('Source path: ' + path)

			# https://docs.python.org/3/using/windows.html#removing-the-max-path-limitation
			# Get the latest modified file from the project folder

			lists = glob.glob(path + r'\*')
			if lists != []:
				latest_file = max(lists, key=os.path.getmtime)
			else:
				print('There is no file under this folder')
				sys.exit()

			# Copy the file to the new folder(source_copy)
			file_name = os.path.basename(latest_file)
			copied_file_path = os.path.join(copied_folder, file_name)
			copyfile(latest_file, copied_file_path)

			self.new_project_location[proj] = copied_file_path

			print('New path: ' + copied_file_path)


		print('Total Missed files', missed_files)
		print('************* Copying project source files: Done')

	def copy_files_to_source(self):
		global missed_files

		for proj in self.new_project_location:
			source_path = self.new_project_location[proj]

			print('Source path: ' + source_path)

			splitnames = []
			file_name, extension = os.path.splitext(os.path.basename(source_path))
			splitnames = file_name.split('_')
			file_name = file_name.replace(str(splitnames[-1]),self.format_time())
			file_name = file_name + extension
			copied_file_path = os.path.join(self.project_location[proj], file_name)
			copyfile(source_path, copied_file_path)
			print('New path: ' + copied_file_path)


		print('Total Missed files', missed_files)
		print('************* Copying source files to original folder: Done')

	def format_time(self):
		t = datetime.datetime.now().timestamp()
		t = str(t)
		a,b = t.split('.')
		return a

	def sync_findings(self, projects):

		for proj in projects:
			wb = load_workbook(self.new_project_location[proj])

			self.update_finding_by_proj(proj, wb)

			wb.save(self.new_project_location[proj])

			self.print_results()


		db.close_connection()

	def sync_findings_all(self, projects):

		for proj in projects:
			wb = load_workbook(self.new_project_location[proj])

			self.update_finding_by_proj(proj, wb)

			wb.save(self.new_project_location[proj])

			self.print_results()
			self.copy_files_to_source()

		# db.close_connection()

	def sync_findings_all_location(self, projects,location):

		for proj in projects:
			wb = load_workbook(self.new_project_location[proj])

			self.update_finding_by_proj(proj, wb)

			wb.save(self.new_project_location[proj])

			self.print_results()

		# db.close_connection()

	def update_finding_by_proj(self, project, workbook):
		print('************* Get findings from project: ' + project)
		actual_sheets = workbook.sheetnames
		print(actual_sheets)

		for sheet in sheets:
			print(sheet)
			if sheet in actual_sheets:
				ws = workbook[sheet]
				finding_df, project_artifact_df = self.get_findings(project, sheet)
				# print(finding_df)
				self.update_findings(project, sheet, finding_df, project_artifact_df,ws)


		# Read Findings for each project
	def get_findings(self, project, sheet):
		
		# Get data for entire sheet. Read once, use elsewhere.
		print(self.new_project_location[project],sheet)
		sheet_df = pd.read_excel(self.new_project_location[project],sheet,header=None)

		# Sheet Data Frame is divided by ROWS: [Header DF] 0:6 and [FINDING DF] 7:onwards
		header_df = sheet_df.iloc[0:header_end,:]
		finding_df = sheet_df.iloc[qa_data_start+1:,0:]

		finding_headers = finding_df.iloc[0]
		finding_headers = finding_headers.tolist()

		if not finding_df.empty:            
			# Prepare the Header.
			finding_df.columns = finding_headers
			finding_df = finding_df[1:]

			# Diff between index and row number
			diff_idx_row = 1
			finding_df['#row'] = finding_df.index + diff_idx_row

		return finding_df, header_df

	# Insert and update findings in database
	def update_findings(self, project, sheet, finding_df, project_artifact_df, ws):
		global count_new, proj_artifacts, findings, last_row
		global pa_error, finding_error
		global count_dict

				# Drop nan columns
		##        project_artifact_df = project_artifact_df.dropna(axis = 1, how='all')

		project_artifact_df = project_artifact_df.iloc[:,2:3]

		project_artifact_df.columns = ['project']

		# project artifact table
		project_artifact_list = project_artifact_df['project'].tolist()

		project_phase = project_artifact_list[1]
		project_mgr = project_artifact_list[2]


		artifact_location = project_artifact_list[4]
		review_by = project_artifact_list[5]
		vendor_name = project_artifact_list[6]

		# new dataframe
		finding_dataframe = pd.DataFrame(columns=finding_df.columns.tolist())
		##    print(finding_df)
		last_row = len(finding_df)

		# read row by row of Finding data
		for index, row in finding_df.iterrows():
			
			try:
				valid_review_row = row[valid_column]
			except:
				continue
			# Valid finding data
			if valid_review_row is not np.nan:
				print('Row: ' + str(row['#row']))
	

				# Project artifact dictionary
				pa_dict = {}
				pa_dict['pa_id'] = -1
				pa_dict['artifact_location'] = artifact_location
				pa_dict['count'] = 0
				if sheet in count_columns_sheets:
					try:
						pa_dict['count'] = row['Test Case Count'] if row['Test Case Count'] is not nan else 0
					except:
						pa_dict['count'] = 0

				pa_dict['project_name'] = project
				pa_dict['project_module'] = row['Project Module'] if row['Project Module'] is not nan else nan_str
				pa_dict['vendor_name'] = vendor_name if vendor_name is not nan else nan_str
				pa_dict['artifact_type'] = sheet
				pa_dict['doc_name_version'] = row['Document Name & Version #'] if row['Document Name & Version #'] is not nan else ''
				pa_dict['review_by'] = review_by


				# Finding artifact dictionary
				find_dict = {}
				find_dict['pa_id'] = -1
				find_dict['artifact'] = sheet
				find_dict['location_in_artifact'] = row['Artifact / Section / Page No.\n'] if row['Artifact / Section / Page No.\n'] is not nan else ''
				find_dict['finding_category'] = row['Finding Type'] if row['Finding Type'] is not nan else nan_str
				qa_review_observation = str(row['QA Comments / Recommendations\n']) if row['QA Comments / Recommendations\n'] is not nan else ''
				project_response = str(row['Project Teams Response / Remediation Date']) if row['Project Teams Response / Remediation Date'] is not nan else ''
				
				# Escape single quote for Access
				find_dict['qa_review_observation'] = qa_review_observation.replace("'","''")
				find_dict['project_response'] = project_response.replace("'","''")

				find_dict['status'] = row['Resolution Status\n (Open / Closed)']
				review_date = row['Review Date']
				print(review_date)

				date = []
				if isinstance(review_date,str):
					review_date = review_date.replace(' ','')
					date = review_date.split(':',1)
					try:
						review_date = datetime.datetime.strptime(date[0],'%m/%d/%Y')
					except:
						review_date = nan_date
				try:
					find_dict['review_date'] = review_date.strftime('%m/%d/%Y')
				except:
					find_dict['review_date'] = nan_date

				find_dict['severity'] = row['Severity'] if row['Severity'] is not nan else nan_str

				followup_1 = row['Follow up Review Date 1'].strftime('%m/%d/%Y') if row['Follow up Review Date 1'] is not nan  and isinstance(row['Follow up Review Date 1'],datetime.date)  else nan_date
				followup_2 = row['Follow up Review Date 2'].strftime('%m/%d/%Y') if row['Follow up Review Date 2'] is not nan  and isinstance(row['Follow up Review Date 2'],datetime.date)  else nan_date
				followup_3 = row['Follow up Review Date 3'].strftime('%m/%d/%Y') if row['Follow up Review Date 3'] is not nan  and isinstance(row['Follow up Review Date 3'],datetime.date)  else nan_date
				followup_4 = row['Follow up Review Date 4'].strftime('%m/%d/%Y') if row['Follow up Review Date 4'] is not nan  and isinstance(row['Follow up Review Date 4'],datetime.date)  else nan_date
				followup_5 = row['Follow up Review Date 5'].strftime('%m/%d/%Y') if row['Follow up Review Date 5'] is not nan  and isinstance(row['Follow up Review Date 5'],datetime.date)  else nan_date
				followup_6 = row['Follow up Review Date 6'].strftime('%m/%d/%Y') if row['Follow up Review Date 6'] is not nan  and isinstance(row['Follow up Review Date 6'],datetime.date)  else nan_date
				followup_7 = row['Follow up Review Date 7'].strftime('%m/%d/%Y') if row['Follow up Review Date 7'] is not nan  and isinstance(row['Follow up Review Date 7'],datetime.date)  else nan_date
				followup_8 = row['Follow up Review Date 8'].strftime('%m/%d/%Y') if row['Follow up Review Date 8'] is not nan  and isinstance(row['Follow up Review Date 8'],datetime.date)  else nan_date
				followup_9 = row['Follow up Review Date 9'].strftime('%m/%d/%Y') if row['Follow up Review Date 9'] is not nan  and isinstance(row['Follow up Review Date 9'],datetime.date)  else nan_date
				followup_10 = row['Follow up Review Date 10'].strftime('%m/%d/%Y') if row['Follow up Review Date 10'] is not nan  and isinstance(row['Follow up Review Date 10'],datetime.date)  else nan_date
				
				find_dict['followup_1'] = followup_1
				find_dict['followup_2'] = followup_2
				find_dict['followup_3'] = followup_3
				find_dict['followup_4'] = followup_4
				find_dict['followup_5'] = followup_5
				find_dict['followup_6'] = followup_6
				find_dict['followup_7'] = followup_7
				find_dict['followup_8'] = followup_8
				find_dict['followup_9'] = followup_9
				find_dict['followup_10'] = followup_10

				# Check if new data or exist
				finding_index = row['finding_id']
				print("Finding index : ", str(finding_index))
				if finding_index is nan:  # New data
					print("new data")
					pa_id = self.insert_project_artifact(pa_dict)
					
					if pa_id != -1:
						find_dict['pa_id'] = pa_id
						print(pa_id)
						proj_artifacts += pa_id,
						finding_id = self.insert_findings(find_dict)

						if finding_id != -1:
							finding_df['finding_id'] = finding_id

							findings += finding_id,

							# Add finding_id in each row
							row['finding_id'] = finding_id
							print(row['finding_id'])
							finding_dataframe = finding_dataframe.append(row)

							count_new += 1
							count_dict[sheet] = count_dict.get(sheet,0) + 1
							# print(row.tolist())
							
						else: # Cannot insert the data in Findings table
							finding_error = str(row['#row'])
							print(row.tolist())
							return
						
					else: # Cannot insert the data in Project Artifact table
						pa_error = str(row['#row'])
						print(row.tolist())
						return

				elif int(finding_index) > 0:   # Existing data: need to be updated
					print("existing data")
					
					find_dict['finding_id'] = finding_index
					pa_dict['finding_id'] = finding_index

					# Update data in database
					self.update_findings_database(find_dict, pa_dict)
					##                print(row.tolist())
				 
		if len(finding_dataframe) > 0:
			self.write_finding_id(ws,finding_dataframe,sheet)
		else:
			if new_project:
				self.create_data_validation(sheet,ws,epy_st,epy_st+10)



	# Write the finding id into excel
	def write_finding_id(self, ws, finding_dataframe,sheet):
		print('********writing finding ids in Project file...')

		first_row = finding_dataframe['#row'].iloc[0]
		
		# Iterate rows and write finding_id
		for index, row in finding_dataframe.iterrows():
			row_num = row['#row']
			row_finding_id = row['finding_id']
			column_num = TestCase_Finding_id_column

			# Update cell value
			excel_cell = ws.cell(row = row_num, column = column_num)
			excel_cell.value = row_finding_id

		# last_row = first_row+len(finding_dataframe)-1

		# create data validation
		if new_project:
			self.create_data_validation(sheet,ws,first_row,last_row)


	# Create data validation
	def create_data_validation(self, sheet,ws,first_row,last_row):
		print('********Creating data validation in Project file...')
		
		observation_col = 'E'
		if sheet not in ['Test Cases','Test Results']:
			observation_col = 'D'

		observation_col = observation_col+ str(first_row)+':'+observation_col+str(last_row)
		print(observation_col)
		print((valid_dict[sheet]))
		# Create a data-validation object with list validation
		dv = DataValidation(type="list", formula1="%s" % (valid_dict[sheet]), allow_blank=True)
		dv.add(observation_col)

		ws.add_data_validation(dv)


	def print_results(self):
		print('**************************End syncing data from excel to Access***********************')
		print(str(count_new) + ' new data added to Access')
		for k in count_dict:
			print(k + ': ' + str(count_dict[k]))

		if proj_artifacts != []:
			print('PA_id: ' + str(proj_artifacts[0]) + ' - ' + str(proj_artifacts[-1]))
		if findings != []:
			print('Findings_id: ' + str(findings[0]) + ' - ' + str(findings[-1]))

		if pa_error != '': 
			print('Errors for Project Artifact: Row ' + pa_error)
		if finding_error != '':
			print('Errors for Finding: Row ' + finding_error)



	'''
		Access Database
	'''
	# Create and update Project Artifact table, return pa_id
	def insert_project_artifact(self, pa_dict):
		
		print('insert_project_artifact')

		return db.insert_pa_access(pa_dict)


	# Create and update Finding table, return finding_id
	def insert_findings(self, find_dict):
		print('********insert_finding************\n')
		return db.insert_finding(find_dict)

	# Update Finding and project artifact table
	def update_findings_database(self, find_dict, pa_dict):
		print('********Updating finding table')

		db.update_project_artifact(pa_dict)
		
		db.update_finding(find_dict)


'''

Report : End
MainWindow : GUI  

'''

class TabOne(Frame):

	def __init__(self, parent):
		Frame.__init__(self, parent)
		self.tab1_note = ttk.Notebook(self,width=parent.winfo_screenwidth(), height=parent.winfo_screenheight())
		tab1_refresh_db = ttk.Frame(self.tab1_note)
		tab1_open_observations = ttk.Frame(self.tab1_note)
		tab1_closed_observations = ttk.Frame(self.tab1_note)

		self.tab1_note.add(tab1_refresh_db, text= "General")
		self.tab1_note.add(tab1_open_observations, text= "Open Projects")
		self.tab1_note.add(tab1_closed_observations, text= "Closed/Deferred Projects")
		self.tab1_note.pack()

		self.tab_one_load(tab1_refresh_db)
		self.tab_two_load(tab1_open_observations)
		self.tab_three_load(tab1_closed_observations)
		widget_list = []

		widget_list.extend(tab1_refresh_db.winfo_children())
		widget_list.extend(tab1_open_observations.winfo_children())
		widget_list.extend(tab1_closed_observations.winfo_children())

		for wid in widget_list:
			try:
				wid.configure(font = 'helvetica 12')
			except:
				pass

	def tab_one_load(self,tab1_refresh_db):

		######################################################################################
		#TAB1
		#TAB1 - TAB1
		######################################################################################
		def sync_tab1():
			global projects
			projects = db.query_projects()
			proj_names = [project[1] for project in projects]
			s = ttk.Style()
			# s.theme_use('clam')
			s.configure("green.Horizontal.TProgressbar", foreground='green', background='green')
			progress_bar = ttk.Progressbar(tab1_refresh_db, style="green.Horizontal.TProgressbar", orient='horizontal',length=300, mode='determinate')
			progress_bar.grid(row=2, column=3,padx=10, pady=10, columnspan=4)
			progress_bar['maximum'] = len(proj_names)
			pval = 1
			progress_bar['value'] = pval
			progress_bar.update_idletasks()
			tt = str(round(pval/len(proj_names) * 100.0,2)) + "%"
			error = []
			for proj in proj_names:
				each_proj = StringVar(self)
				each_proj.set(proj)
				print ("Project selected:", each_proj.get())
				project_name = each_proj.get()
				try:
					Report([project_name])
					pval = pval + 1 ## Or whatever increment makes sense.
					progress_bar['value'] = pval
				except:
					print("Error in uploading...")
					error.append(project_name)
					continue

				progress_bar.update_idletasks()
				tt = str(round(pval/len(proj_names) * 100.0,2)) + "%"
				label = Label(tab1_refresh_db, text=tt, font=("Helvetica", 12),fg="green")
				label.grid(row=2,column=6,padx=10,pady=5)
				tab1_refresh_db.after(1500, label.destroy)

			label = Label(tab1_refresh_db, text=tt, font=("Helvetica", 12),fg="green")
			label.grid(row=2,column=6,padx=10,pady=5)
			
			if error == []:
				tt = 'No errors in database'
				label_error = Label(tab1_refresh_db, text=tt,font=("Helvetica", 12),fg="green")
				label_error.grid(row=0, column=0, columnspan=2, rowspan=2, sticky=W+E+N+S, padx=5, pady=5)
				tab1_refresh_db.after(500, label_error.destroy)
				return
			
			error_label = Label(tab1_refresh_db, text="Upload Errors")
			error_label.grid(row=2, column=7,padx=20, pady=10)
			
			listbox_error_select = StringVar(tab1_refresh_db,value='Default Text')
			list_box_error = Listbox(tab1_refresh_db,listvariable=listbox_error_select,selectmode='browse',width=60,height=4)
			listbox_error_select.set(error)
			list_box_error.grid(row=2, column=8,padx=20,pady=10, rowspan=4)
			# list_box_open.bind("<Double-Button-1>", list_has_changed)


		lbl_refresh_title = Label(tab1_refresh_db, text='Refresh Database?')
		lbl_refresh_title.grid(row=0, column = 0,padx=10, pady=20)

		# now = datetime.datetime.now()
		# date_display = now.strftime('%A, %d %B %Y, %H:%M.')
		lbl = Label(tab1_refresh_db, text="CITSQ Productivity Tool")
		lbl.grid(row=0, column = 2,padx=10, pady=20)
		
		lbl_refresh = Label(tab1_refresh_db, text="Refresh") 
		lbl_refresh.grid(row=2, column=0,padx=10, pady=10) 
		button_refersh = Button(tab1_refresh_db, text="Update Database", command=sync_tab1)
		button_refersh.grid(row =2, column =2,padx=10, pady=10)

		lbl_last_updated = Label(tab1_refresh_db, text="Current Login By: ") 
		lbl_last_updated.grid(row=4, column=0,padx=10, pady=10) 
		lbl_last_updated_name = Label(tab1_refresh_db, text=str(username)) 
		lbl_last_updated_name.grid(row=4, column=2,padx=10, pady=10)


		# f=Frame(tab1_refresh_db,height=1,width=300,bg="black")
		# f.grid(row=4,padx=10, pady=30) 

		# lbl_refresh_prod = Label(tab1_refresh_db, text='Projects in Production',font=11)
		# lbl_refresh_prod.grid(row=5, column = 2,padx=10, pady=10)

		# lf = ttk.Labelframe(tab1_refresh_db, text='', labelanchor=N+S, borderwidth=2,relief="sunken",width=400)
		# lf.grid(row=6, column = 2,padx=10, pady=10)

		# for i in range(1,6):
		#     lbl_last_updated = Label(lf, text=" Project " + str(i)) 
		#     lbl_last_updated.grid(row=i+5, column=0,padx=10, pady=10) 

	def tab_two_load(self,tab1_open_observations):
		######################################################################################
		#TAB2
		#TAB2 - TAB2
		######################################################################################
		projects = db.query_projects();
		current_status = [1]
		fig_artif = Figure(figsize=(5,3))
		fig_findin = Figure(figsize=(6,3))
		ax = fig_artif.add_subplot(111)
		ay = fig_findin.add_subplot(111)

		listbox_observation_select = StringVar(tab1_open_observations,value='Default Text')
		list_box_open = Listbox(tab1_open_observations,listvariable=listbox_observation_select,selectmode='browse',width=60,height=4)

		canvas_artifact = FigureCanvasTkAgg(fig_artif, master=tab1_open_observations)
		canvas_artifact.get_tk_widget().grid(row=7,column=0,padx=10,pady=5,columnspan=3)

		canvas_finding = FigureCanvasTkAgg(fig_findin, master=tab1_open_observations)
		canvas_finding.get_tk_widget().grid(row=7,column=4,padx=10,pady=5,columnspan=3)

		def display_record(record):
			a,b,c,d,e,f,g,h,i,j,k,l,m,n,o = record

			#Header
			a_doc_name_and_ver = a
			b_test_count = b
			c_project_name = c
			d_module_name = d
			f_finding_type = f
			o_artifact_type = o
			n_finding_id = n

			#Review
			e_location_in_artifact = e
			g_QA_review = g
			i_review_date = i

			#Response
			h_project_response = h
			follow_up_reviews = [j,k,l,m]

			follow_up_text = ''
			for idx,date_text in enumerate(follow_up_reviews):
				if(date_text <= default_date):
					follow_up_reviews[idx] = ' '
				else:
					follow_up_text = follow_up_text + 'Follow Up %d:\t%s\n' % ((idx+1),follow_up_reviews[idx].strftime('%b %d %Y %I:%M%p'))

			header_text = "Finding ID:\n%s\n\nDoc Name & Version:\n%s\n\nTest Count:\n%s\n\nProject Name:\n%s\n\nModule Name:\n%s\n\nFinding Type:\n%s\n\nArtifact Name:\n%s\n\n" % (n,a,b,c,d,f,o)
			review_text = "\nArtifact Location:\n%s\n\nQA Review:\n%s\n\nReview Date:\n%s\n" % (e,g,i.strftime('%b %d %Y %I:%M %p'))
			response_text = "\nProject Response:\n%s\n\n" % (h)
			response_text = response_text + follow_up_text

			full_text = header_text + review_text + response_text
			width, height = 60,1
			text_response = scrolledtext.ScrolledText(tab1_open_observations,width=width, height=height, wrap='word')
			text_response.insert(1.0, full_text)
			text_response['font'] = ('consolas', '10','bold')
			text_response.grid(row=2,column=4,padx=20,pady=10,rowspan=4,sticky="news")

		def list_has_changed(self):
			### CHANGE THIS IF SELECTMODE IS MULTIPLE
			values = [list_box_open.get(idx) for idx in list_box_open.curselection()]
			if values != []:
				item = values[0]
				a,b,c,d = item.split("  ->  ")
				b = int(b)
				query_result = db.query_findings_finding_id(b)
				if(query_result != ''):
					display_record(query_result[0])

		def update_modules(a,b,c):
			# 2 # # Current project modules
			proj_name = proj_select.get()

			query_projects_location = db.query_projects_location(proj_name);
			proj_location_select.set(query_projects_location)

			project_modules = db.query_project_modules(proj_name);
			proj_mo_names = []
			if project_modules == []:
				print('No project modules in database for the selected project')
				label = Label(tab1_open_observations, text="No Modules in this project!")
				label.grid(row=0, column=0, sticky=W+E+N+S, padx=5, pady=5)
				tab1_open_observations.after(500, label.destroy)
				proj_mod_select.set("No Modules")
				project_module_dropdown = OptionMenu(tab1_open_observations,proj_mod_select,[])

			else:
				proj_mo_names = [module[0] for module in project_modules]
				proj_mod_select.set(proj_mo_names[0])
				project_module_dropdown = OptionMenu(tab1_open_observations,proj_mod_select,*proj_mo_names)

			# print(project_modules)
			project_module_dropdown.configure(font='helvetica 12')
			project_module_dropdown.grid(row=2, column=1,padx=10, pady=10)

		def update_teams(a,b,c):
			teams = db.query_teams()
			if(teams == []):
				print('No Teams found')
				return

			team_list = [team[0] for team in teams]
			team_list.insert(0,'All')
			team_select.set(team_list[0])
			team_dropdown = OptionMenu(tab1_open_observations,team_select,*team_list)
			team_dropdown.configure(font='helvetica 12')
			team_dropdown.grid(row=3,column=1,padx=10,pady=10)

		def update_artifacts(a,b,c):
			team_name,proj_name,mod_name = team_select.get(),proj_select.get(),proj_mod_select.get()

			if(team_name == 'All'):
				team_name = db.query_team_name()
			else:
				team_name = [team_name]

			artifacts_fr_project = db.query_artifacts_details(team_name,proj_name,mod_name,current_status)
			title_text = proj_name + " - " + mod_name
			
			if artifacts_fr_project == []:
				print('No artifacts in database')
				label_error = Label(tab1_open_observations, text="No artifacts in the database!")
				label_error.grid(row=0, column=0, columnspan=2, rowspan=2, sticky=W+E+N+S, padx=5, pady=5)
				tab1_open_observations.after(1500, label_error.destroy)
				return 

			artifact_names = [artifact[0] for artifact in artifacts_fr_project]
			if artifact_names != []:
				artifact_names.insert(0,'All')
				artifact_type_select.set(artifact_names[0])
				artifact_type_dropdown = OptionMenu(tab1_open_observations,artifact_type_select,*artifact_names)
				artifact_type_dropdown.configure(font='helvetica 12')
				artifact_type_dropdown.grid(row=4, column=1,padx=10, pady=10)

		def update_findings(a,b,c):
			artifact_name,team_name,proj_name,mod_name = artifact_type_select.get(),team_select.get(),proj_select.get(),proj_mod_select.get()
			if(artifact_name == 'All'):
				artifact_name = db.query_artifact_name()
			else:
				artifact_name = [artifact_name]

			if(team_name == 'All'):
				team_name = db.query_team_name()
			else:
				team_name = [team_name]

			observations_fr_project = db.query_findings(team_name,proj_name,mod_name,artifact_name,current_status)
			
			if observations_fr_project == []:
				print('No findings in database')
				label_error = Label(tab1_open_observations, text="No findings in the database!")
				label_error.grid(row=0, column=0, columnspan=2, rowspan=2, sticky=W+E+N+S, padx=5, pady=5)
				tab1_open_observations.after(500, label_error.destroy)
				return 

			# max_len = 40
			finding_names = [finding[0] for finding in observations_fr_project]
			# finding_names = [finding[0][:max_len] + '..' if len(finding[0]) > max_len else finding[0] for finding in observations_fr_project ]

			if finding_names != []:
				finding_names.insert(0,'All')
				finding_type_select.set(finding_names[0])
				finding_type_dropdown = OptionMenu(tab1_open_observations,finding_type_select,*finding_names)
				finding_type_dropdown.configure(font='helvetica 12')
				finding_type_dropdown.grid(row=5, column=1,padx=10, pady=10)
				tab1_open_observations.grid_columnconfigure(1, weight=1)
				tab1_open_observations.grid_columnconfigure(4, weight=1)

		def update_listbox(a,b,c):
			finding_name,artifact_name,team_name,proj_name,mod_name = finding_type_select.get(),artifact_type_select.get(),team_select.get(),proj_select.get(),proj_mod_select.get()
			if(artifact_name == 'All'):
				artifact_name = db.query_artifact_name()
			else:
				artifact_name = [artifact_name]

			if(finding_name == 'All'):
				finding_name = db.query_finding_name()
			else:
				finding_name = [finding_name]

			if(team_name == 'All'):
				team_name = db.query_team_name()
			else:
				team_name = [team_name]

			listbox_for_findings = db.query_findings_details(team_name,proj_name,mod_name,artifact_name,finding_name,current_status)
			if listbox_for_findings == []:
				print('No projects in database')
				label_error = Label(tab1_open_observations, text="No Findings in the database!")
				label_error.grid(row=0, column=0, columnspan=2, rowspan=2, sticky=W+E+N+S, padx=5, pady=5)
				tab1_open_observations.after(500, label_error.destroy)
				return

			findings_list = [str(find[1]) + "  ->  " + str(find[0]) + "  ->  " + str(find[4]) + "  ->  " + str(find[2]) for find in listbox_for_findings]
			listbox_observation_select.set(findings_list)
			list_box_open.grid(row=0, column=4,padx=20,pady=10, rowspan=2)
			list_box_open.bind("<Double-Button-1>", list_has_changed)
		
		def animate_artifact_graph(i):
			#######################################
			''' GRAPH START '''
			#######################################
			## DATA
			x_cat = [x[1] for x in db.query_artifact()]
			x_cat = x_cat[:-2]


			y_count = {}
			for i in x_cat:
				y_count[i] = 0

			artifact_name,proj_name,mod_name,team_name = artifact_type_select.get() ,proj_select.get(),proj_mod_select.get(),team_select.get()
			if(artifact_name == 'All'):
				artifact_name = db.query_artifact_name()
			else:
				artifact_name = [artifact_name]

			if(team_name == 'All'):
				team_name = db.query_team_name()
			else:
				team_name = [team_name]

			observations_fr_project = db.query_findings_by_projects(team_name,proj_name,mod_name,artifact_name,current_status)

			if(observations_fr_project == ''):
				ax.clear()
				tt = "No project modules in database for the selected project"
				label = Label(tab1_open_observations, text=tt, font=("Helvetica", 12))
				label.grid(row=7,column=0,padx=10,pady=5,columnspan=3)
				tab1_open_observations.after(500, label.destroy)
				return

			title = proj_name + " - " + mod_name

			for i in observations_fr_project:
				y_count[i[2]] = y_count[i[2]] + 1

			y_count = { k:v for k, v in y_count.items() if v > 0 and k != 'RTM'}

			x = list(y_count.keys())
			y = list(y_count.values())

			ax.clear()
			now = datetime.datetime.now()
			date_display = now.strftime('%A, %d %B %Y, %H:%M')
			ax.set_title (title + "\n" + "Total number of findings per deliverable type\n(as of " + date_display + ")", fontsize=8)

			def func(pct, allvals):
				absolute = int(pct/100.*np.sum(allvals))
				return "{:.1f}%\n{:d}".format(pct, absolute)


			wedges, texts, autotexts = ax.pie(y, autopct=lambda pct: func(pct, y),
											  textprops=dict(color="w"))

			ax.legend(wedges, x,
					  title="Artifact Types",
					  loc="center left",
					  bbox_to_anchor=(1, 0, 0.5, 1))
			fig_artif.tight_layout()
			ax.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.

			#######################################
			''' GRAPH END '''
			#######################################

		def animate_finding_graph(i):
			#######################################
			''' GRAPH START '''
			#######################################

			## DATA
			x_cat = [x[1] for x in db.query_finding()]
			y_count = {}

			for i in x_cat:
				y_count[i] = 0

			artifact_name,proj_name,mod_name,team_name = artifact_type_select.get() ,proj_select.get(),proj_mod_select.get(),team_select.get()
			artif_name = artifact_name
			tim_name = team_name

			if(artifact_name == 'All'):
				artifact_name = db.query_artifact_name()
				artif_name = 'All Artifacts'
			else:
				artif_name = artifact_name
				artifact_name = [artifact_name]


			if(team_name == 'All'):
				team_name = db.query_team_name()
			else:
				team_name = [team_name]

			observations_fr_project = db.query_findings_by_projects(team_name,proj_name,mod_name,artifact_name,current_status)
			if(observations_fr_project == ''):
				ay.clear()
				tt = "No project modules in database for the selected project"
				label = Label(tab1_open_observations, text=tt, font=("Helvetica", 12))
				label.grid(row=7,column=4,padx=10,pady=5,columnspan=3)
				tab1_open_observations.after(500, label.destroy)
				return
			title = proj_name + " - " + mod_name

			for i in observations_fr_project:
				y_count[i[1]] = y_count[i[1]] + 1

			y_count = { k:v for k, v in y_count.items() if v > 0 and k != 'nan'}
			y_count = OrderedDict(sorted(y_count.items(), key=lambda kv: kv[1],reverse=True))

			if(len(list(y_count.keys())) >= 5):
				to_remove = list(y_count.keys())[5:]
				for x in to_remove:
					del y_count[x]

			x = list(y_count.values())
			y = list(y_count.keys())

			## SHOW
			ay.clear()
			bar_width = 0.4
			ay.barh(y,x,bar_width,color='yellow')
			ay.invert_yaxis()
			rects = ay.patches
			# print("rects",len(rects))
			labels = [ i for i in list(y_count.values())]
			# print("\n")

			for rect, label in zip(rects, labels):
				height = rect.get_height()/2
				width = rect.get_width() - 0.50
				ay.text(rect.get_x() + width, rect.get_y()+height,label,fontsize=8)

			ay.tick_params(
			axis='x',          # changes apply to the x-axis
			which='both',      # both major and minor ticks are affected
			bottom=False,      # ticks along the bottom edge are off
			top=False,         # ticks along the top edge are off
			labelbottom=False) # labels along the bottom edge are off

			ay.set_yticklabels(y,fontsize=6,wrap=True)

			for tick in ay.yaxis.get_major_ticks():
				tick.label1.set_verticalalignment('center')

			ay.set_title (title + "\n" + "Top 5 Deliverables (" + str(artif_name) +")", fontsize=8)

			#######################################
			## ''' GRAPH END '''
			#######################################


		def artifact_graph():
			# canvas_artifact = FigureCanvasTkAgg(fig_artif, master=tab1_open_observations)
			# canvas_artifact.get_tk_widget().grid(row=7,column=0,padx=10,pady=5,columnspan=3)
			canvas_artifact.draw()  

			def file_save():
				animate_artifact_graph(0)
				canvas_artifact.draw() 
				orig_color = save_as_artifact_graph.cget("background")

				def change(orig_color):
					save_as_artifact_graph.configure(background = orig_color)

				filename =  filedialog.asksaveasfilename(initialdir = "/",title = "Select file",filetypes = (("PNG files","*.png"),("JPEG files","*.jpg"),("all files","*.*")))
				if os.path.basename(filename) is '':
					save_as_artifact_graph.configure(background = "red")
					tab1_open_observations.after(1000, lambda: change(orig_color))
					return
				try:
					fig_artif.savefig(filename)
				except:
					save_as_artifact_graph.configure(background = "red")
					tab1_open_observations.after(1000, lambda: change(orig_color))
					return

				save_as_artifact_graph.configure(background = "green")
				tt = "Exported"
				label = Label(tab1_open_observations, text=tt, font=("Helvetica", 12))
				label.grid(row=0,column=0,padx=10,pady=5,columnspan=3)

				tab1_open_observations.after(500, lambda: change(orig_color))
				tab1_open_observations.after(500, label.destroy)
				print(filename)

			def export_win():
				output_files = current_output_folder + "/Open/" + proj_select.get() + " -- " + proj_mod_select.get()

				animate_artifact_graph(0)
				canvas_artifact.draw() 

				orig_color = export_artifact_graph.cget("background")

				def change(orig_color):
					export_artifact_graph.configure(background = orig_color)

				if not os.path.exists(output_files):
					os.makedirs(output_files)
				try:
					fig_artif.savefig(output_files + "/artifact.png")
				except:
					export_artifact_graph.configure(background = "red")
					tab1_open_observations.after(1000, lambda: change(orig_color))
					return

				export_artifact_graph.configure(background = "green")
				tt = "Exported"
				label = Label(tab1_open_observations, text=tt, font=("Helvetica", 12))
				label.grid(row=0,column=0,padx=10,pady=5,columnspan=3)

				tab1_open_observations.after(500, lambda: change(orig_color))
				tab1_open_observations.after(500, label.destroy)

			export_artifact_graph = Button(tab1_open_observations, text='Export', command=export_win)
			export_artifact_graph.grid(row=6,column=0,padx=30,pady=20,sticky='we')

			save_as_artifact_graph = Button(tab1_open_observations, text='Save As', command=file_save)
			save_as_artifact_graph.grid(row=6,column=2,padx=5,pady=20,sticky='we')

		def finding_category():
			# canvas_finding = FigureCanvasTkAgg(fig_findin, master=tab1_open_observations)
			# canvas_finding.get_tk_widget().grid(row=7,column=4,padx=10,pady=5,columnspan=3)
			canvas_finding.draw() 

			def file_save():
				animate_finding_graph(0)
				canvas_finding.draw() 
				orig_color = save_as_finding_graph.cget("background")

				def change(orig_color):
					save_as_finding_graph.configure(background = orig_color)

				filename =  filedialog.asksaveasfilename(initialdir = "/",title = "Select file",filetypes = (("PNG files","*.png"),("JPEG files","*.jpg"),("all files","*.*")))
				if os.path.basename(filename) is '':
					save_as_artifact_graph.configure(background = "red")
					tab1_open_observations.after(1000, lambda: change(orig_color))
					return
				try:
					fig_findin.savefig(filename)
				except:
					save_as_finding_graph.configure(background = "red")
					tab1_open_observations.after(1000, lambda: change(orig_color))
					return

				save_as_finding_graph.configure(background = "green")
				tt = "Exported"
				label = Label(tab1_open_observations, text=tt, font=("Helvetica", 12))
				label.grid(row=0,column=0,padx=10,pady=5,columnspan=3)
				tab1_open_observations.after(500, lambda: change(orig_color))
				tab1_open_observations.after(500, label.destroy)
				print(filename)

			def export_win():
				output_files = current_output_folder + "/Open/" + proj_select.get() + " -- " + proj_mod_select.get()

				animate_finding_graph(0)
				canvas_finding.draw() 

				orig_color = export_finding_graph.cget("background")
				def change(orig_color):
					export_finding_graph.configure(background = orig_color)

				if not os.path.exists(output_files):
					os.makedirs(output_files)

				try:
					fig_findin.savefig(output_files + "/finding.png")
				except:
					export_finding_graph.configure(background = "red")
					tab1_open_observations.after(1000, lambda: change(orig_color))
					return

				export_finding_graph.configure(background = "green")
				tt = "Exported"
				label = Label(tab1_open_observations, text=tt, font=("Helvetica", 12))
				label.grid(row=0,column=0,padx=10,pady=5,columnspan=3)

				tab1_open_observations.after(500, lambda: change(orig_color))
				tab1_open_observations.after(500, label.destroy)

			export_finding_graph = Button(tab1_open_observations, text='Export', command=export_win)
			export_finding_graph.grid(row=6,column=4,padx=30,pady=20,sticky='we')

			save_as_finding_graph = Button(tab1_open_observations, text='Save As', command=file_save)
			save_as_finding_graph.grid(row=6,column=5,padx=5,pady=20,sticky='we')

		def update_graphs(a,b,c):
			animate_artifact_graph(0)
			animate_finding_graph(0)
			canvas_artifact.draw()  
			canvas_finding.draw() 
			# # button_update_graphs.config(text="AUTO REFRESH is ON")

		def refresh_projects():
			projects = db.query_projects();
			if(projects == []):
			    print('No projects in database')
			    return

			proj_names = [project[1] for project in projects]
			proj_select.set(proj_names[0])
			project_dropdown = OptionMenu(tab1_open_observations,proj_select,*proj_names)
			project_dropdown.configure(font='helvetica 12')
			project_dropdown.grid(row=1,column=1,padx=10,pady=10)

		def feedback_report():
			print(proj_location_select.get())
			lists = glob.glob(proj_location_select.get() + r'\*')
			if lists != []:
				latest_file = max(lists, key=os.path.getmtime)
			else:
				tt = 'There is no file under this folder'
				label = Label(tab1_open_observations, text=tt, font=("Helvetica", 12), fg="red")
				label.grid(row=0,column=0,padx=10,pady=5)
				tab1_open_observations.after(1000, label.destroy)
				return
			file_name = os.path.basename(latest_file)
			print(latest_file)
			copied_file_path = os.path.join(copied_folder, file_name)
			copyfile(latest_file, copied_file_path)
			os.system('start excel \"%s\"' % (copied_file_path))

			# self.new_project_location[proj] = copied_file_path

			file = filedialog.askopenfile(parent=tab1_open_observations,initialdir=proj_location_select.get(),filetypes=[("Excel Files", "*.xlsx")],mode='rb',title='Open Feedback Reports')
			if file != None:
				data = file.read()
				file.close()
				print("I got %d bytes from this file.", len(data))

		#Title
		closed_caption = Label(tab1_open_observations, text='Open Observations')
		closed_caption.grid(row=0, column = 0,padx=10, pady=10)

		#Data Labels
		project_name_label = Label(tab1_open_observations, text="Project Name")
		project_module_label = Label(tab1_open_observations, text="Current Project Modules") 
		team_label = Label(tab1_open_observations, text="Team Name")

		artifact_type_label = Label(tab1_open_observations, text="Artifact Types") 
		finding_type_label = Label(tab1_open_observations, text="Finding Category") 
		finding_label = Label(tab1_open_observations, text="Findings")
		finding_detail_label = Label(tab1_open_observations, text="Finding Details") 

		project_name_label.grid(row=1, column=0,padx=10, pady=10) 
		project_module_label.grid(row=2, column=0,padx=10, pady=10) 
		team_label.grid(row=3, column=0,padx=10, pady=10) 
		artifact_type_label.grid(row=4, column=0,padx=10, pady=10)
		finding_type_label.grid(row=5, column=0,padx=10, pady=10)
		finding_label.grid(row=0, column=3,padx=40, pady=10)
		finding_detail_label.grid(row=2, column=3,padx=40, pady=10)

		#String Vars Hidden Fields
		proj_select = StringVar(tab1_open_observations,value='Default Text')
		proj_select.trace('w',update_modules)
		proj_select.trace('w',update_teams)
		proj_select.trace('w',update_artifacts)
		proj_select.trace('w',update_findings)
		proj_select.trace('w',update_listbox)
		proj_select.trace('w',update_graphs)

		proj_mod_select = StringVar(tab1_open_observations,value='Default Text')
		proj_mod_select.trace('w',update_teams)
		proj_mod_select.trace('w',update_artifacts)
		proj_mod_select.trace('w',update_findings)
		proj_mod_select.trace('w',update_listbox)
		proj_mod_select.trace('w',update_graphs)

		team_select = StringVar(tab1_open_observations,value='Default Text')
		team_select.trace('w',update_artifacts)
		team_select.trace('w',update_findings)
		team_select.trace('w',update_listbox)
		team_select.trace('w',update_graphs)

		artifact_type_select = StringVar(tab1_open_observations,value='Default Text')
		artifact_type_select.trace('w',update_findings)
		artifact_type_select.trace('w',update_listbox)
		artifact_type_select.trace('w',update_graphs)

		finding_type_select = StringVar(tab1_open_observations,value='Default Text')
		finding_type_select.trace('w',update_listbox)
		finding_type_select.trace('w',update_graphs)

		proj_location_select = StringVar(tab1_open_observations,value='Default Text')

		#Dropdowns
		if(projects == []):
			print('No projects in database')
			return

		proj_names = [project[1] for project in projects]
		proj_select.set(proj_names[0])
		project_dropdown = OptionMenu(tab1_open_observations,proj_select,*proj_names)
		project_dropdown.grid(row=1,column=1,padx=10,pady=10)

		artifact_graph()
		finding_category()

		button_update_graphs = Button(tab1_open_observations, text="Refresh Projects", command=refresh_projects)
		button_update_graphs.grid(row=5, column=3,padx=40, pady=10)

		button_update_graphs = Button(tab1_open_observations, text="Refresh Graphs", command=lambda:update_graphs(0,0,0))
		button_update_graphs.grid(row=6, column=3,padx=40, pady=10)

		button_open_feedback = Button(tab1_open_observations, text="Feedback Report", command=feedback_report)
		button_open_feedback.grid(row=4, column=3,padx=20, pady=20)   
		button_open_feedback.bind()


	def tab_three_load(self,tab1_closed_observations):
		######################################################################################
		#TAB2
		#TAB2 - TAB2
		######################################################################################
		current_status = [2,3]

		fig_artif = Figure(figsize=(5,3))
		fig_findin = Figure(figsize=(6,3))

		ax = fig_artif.add_subplot(111)
		ay = fig_findin.add_subplot(111)

		listbox_observation_select = StringVar(tab1_closed_observations,value='Default Text')
		list_box_closed = Listbox(tab1_closed_observations,listvariable=listbox_observation_select,selectmode='browse',width=60,height=4)

		canvas_artifact = FigureCanvasTkAgg(fig_artif, master=tab1_closed_observations)
		canvas_artifact.get_tk_widget().grid(row=7,column=0,padx=10,pady=5,columnspan=3)

		canvas_finding = FigureCanvasTkAgg(fig_findin, master=tab1_closed_observations)
		canvas_finding.get_tk_widget().grid(row=7,column=4,padx=10,pady=5,columnspan=3)

		projects = db.query_projects()

		def display_record(record):
			a,b,c,d,e,f,g,h,i,j,k,l,m,n,o = record

			#Header
			a_doc_name_and_ver = a
			b_test_count = b
			c_project_name = c
			d_module_name = d
			f_finding_type = f
			o_artifact_type = o
			n_finding_id = n

			#Review
			e_location_in_artifact = e
			g_QA_review = g
			i_review_date = i

			#Response
			h_project_response = h
			follow_up_reviews = [j,k,l,m]

			follow_up_text = ''
			for idx,date_text in enumerate(follow_up_reviews):
				if(date_text <= default_date):
					follow_up_reviews[idx] = ' '
				else:
					follow_up_text = follow_up_text + 'Follow Up %d:\t%s\n' % ((idx+1),follow_up_reviews[idx].strftime('%b %d %Y %I:%M%p'))

			header_text = "Finding ID:\n%s\n\nDoc Name & Version:\n%s\n\nTest Count:\n%s\n\nProject Name:\n%s\n\nModule Name:\n%s\n\nFinding Type:\n%s\n\nArtifact Name:\n%s\n\n" % (n,a,b,c,d,f,o)
			review_text = "\nArtifact Location:\n%s\n\nQA Review:\n%s\n\nReview Date:\n%s\n" % (e,g,i.strftime('%b %d %Y %I:%M %p'))
			response_text = "\nProject Response:\n%s\n\n" % (h)
			response_text = response_text + follow_up_text

			full_text = header_text + review_text + response_text
			width, height = 60,1
			text_response = scrolledtext.ScrolledText(tab1_closed_observations,width=width, height=height, wrap='word')
			text_response.insert(1.0, full_text)
			text_response['font'] = ('consolas', '10','bold')
			text_response.grid(row=2,column=4,padx=20,pady=10,rowspan=4,sticky="news")

		def list_has_changed(self):
			### CHANGE THIS IF SELECTMODE IS MULTIPLE
			values = [list_box_closed.get(idx) for idx in list_box_closed.curselection()]
			if values != []:
				item = values[0]
				a,b,c,d = item.split("  ->  ")
				b = int(b)
				query_result = db.query_findings_finding_id(b)
				if(query_result != ''):
					display_record(query_result[0])

		def update_modules(a,b,c):
			# 2 # # Current project modules
			proj_name = proj_select.get()

			query_projects_location = db.query_projects_location(proj_name);
			proj_location_select.set(query_projects_location)

			project_modules = db.query_project_modules(proj_name);
			if project_modules == []:
				print('No project modules in database for the selected project')
				label = Label(tab1_closed_observations, text="No Modules in this project!")
				label.grid(row=0, column=0, sticky=W+E+N+S, padx=5, pady=5)
				tab1_closed_observations.after(500, label.destroy)
				proj_mod_select.set("No Modules")
				project_module_dropdown = OptionMenu(tab1_closed_observations,proj_mod_select,[])

			else:
				proj_mo_names = [module[0] for module in project_modules]
				proj_mod_select.set(proj_mo_names[0])
				project_module_dropdown = OptionMenu(tab1_closed_observations,proj_mod_select,*proj_mo_names)

			# print(project_modules)
			project_module_dropdown.configure(font='helvetica 12')
			project_module_dropdown.grid(row=2, column=1,padx=10, pady=10)
		
		def update_teams(a,b,c):
			teams = db.query_teams()
			if(teams == []):
				print('No Teams found')
				return

			team_list = [team[0] for team in teams]
			team_list.insert(0,'All')
			team_select.set(team_list[0])
			team_dropdown = OptionMenu(tab1_closed_observations,team_select,*team_list)
			team_dropdown.configure(font='helvetica 12')
			team_dropdown.grid(row=3,column=1,padx=10,pady=10)

		def update_artifacts(a,b,c):
			team_name,proj_name,mod_name = team_select.get(),proj_select.get(),proj_mod_select.get()

			if(team_name == 'All'):
				team_name = db.query_team_name()
			else:
				team_name = [team_name]

			artifacts_fr_project = db.query_artifacts_details(team_name,proj_name,mod_name,current_status)
			title_text = proj_name + " - " + mod_name
			
			if artifacts_fr_project == []:
				print('No artifacts in database')
				label_error = Label(tab1_closed_observations, text="No artifacts in the database!")
				label_error.grid(row=0, column=0, columnspan=2, rowspan=2, sticky=W+E+N+S, padx=5, pady=5)
				tab1_closed_observations.after(1500, label_error.destroy)
				return 

			artifact_names = [artifact[0] for artifact in artifacts_fr_project]
			if artifact_names != []:
				artifact_names.insert(0,'All')
				artifact_type_select.set(artifact_names[0])
				artifact_type_dropdown = OptionMenu(tab1_closed_observations,artifact_type_select,*artifact_names)
				artifact_type_dropdown.configure(font='helvetica 12')
				artifact_type_dropdown.grid(row=4, column=1,padx=10, pady=10)

		def update_findings(a,b,c):
			artifact_name,team_name,proj_name,mod_name = artifact_type_select.get(),team_select.get(),proj_select.get(),proj_mod_select.get()
			if(artifact_name == 'All'):
				artifact_name = db.query_artifact_name()
			else:
				artifact_name = [artifact_name]

			if(team_name == 'All'):
				team_name = db.query_team_name()
			else:
				team_name = [team_name]

			observations_fr_project = db.query_findings(team_name,proj_name,mod_name,artifact_name,current_status)
			
			if observations_fr_project == []:
				print('No findings in database')
				label_error = Label(tab1_closed_observations, text="No findings in the database!")
				label_error.grid(row=0, column=0, columnspan=2, rowspan=2, sticky=W+E+N+S, padx=5, pady=5)
				tab1_closed_observations.after(500, label_error.destroy)
				return 

			# max_len = 40
			finding_names = [finding[0] for finding in observations_fr_project]
			# finding_names = [finding[0][:max_len] + '..' if len(finding[0]) > max_len else finding[0] for finding in observations_fr_project ]

			if finding_names != []:
				finding_names.insert(0,'All')
				finding_type_select.set(finding_names[0])
				finding_type_dropdown = OptionMenu(tab1_closed_observations,finding_type_select,*finding_names)
				finding_type_dropdown.configure(font='helvetica 12')
				finding_type_dropdown.grid(row=5, column=1,padx=10, pady=10)
				tab1_closed_observations.grid_columnconfigure(1, weight=1)
				tab1_closed_observations.grid_columnconfigure(4, weight=1)

		def update_listbox(a,b,c):
			finding_name,artifact_name,team_name,proj_name,mod_name = finding_type_select.get(),artifact_type_select.get(),team_select.get(),proj_select.get(),proj_mod_select.get()
			if(artifact_name == 'All'):
				artifact_name = db.query_artifact_name()
			else:
				artifact_name = [artifact_name]

			if(finding_name == 'All'):
				finding_name = db.query_finding_name()
			else:
				finding_name = [finding_name]

			if(team_name == 'All'):
				team_name = db.query_team_name()
			else:
				team_name = [team_name]

			listbox_for_findings = db.query_findings_details(team_name,proj_name,mod_name,artifact_name,finding_name,current_status)
			if listbox_for_findings == []:
				print('No projects in database')
				label_error = Label(tab1_closed_observations, text="No Findings in the database!")
				label_error.grid(row=0, column=0, columnspan=2, rowspan=2, sticky=W+E+N+S, padx=5, pady=5)
				tab1_closed_observations.after(500, label_error.destroy)
				return

			findings_list = [str(find[1]) + "  ->  " + str(find[0]) + "  ->  " + str(find[4]) + "  ->  " + str(find[2]) for find in listbox_for_findings]
			listbox_observation_select.set(findings_list)
			list_box_closed.grid(row=0, column=4,padx=20,pady=10, rowspan=2)
			list_box_closed.bind("<Double-Button-1>", list_has_changed)
		
		def animate_artifact_graph(i):
			#######################################
			''' GRAPH START '''
			#######################################
			## DATA
			x_cat = [x[1] for x in db.query_artifact()]
			x_cat = x_cat[:-2]


			y_count = {}
			for i in x_cat:
				y_count[i] = 0

			artifact_name,proj_name,mod_name,team_name = artifact_type_select.get() ,proj_select.get(),proj_mod_select.get(),team_select.get()
			if(artifact_name == 'All'):
				artifact_name = db.query_artifact_name()
			else:
				artifact_name = [artifact_name]

			if(team_name == 'All'):
				team_name = db.query_team_name()
			else:
				team_name = [team_name]

			observations_fr_project = db.query_findings_by_projects(team_name,proj_name,mod_name,artifact_name,current_status)
			print(observations_fr_project)
			if(observations_fr_project == ''):
				ax.clear()
				tt = "No project modules in database for the selected project"
				label = Label(tab1_closed_observations, text=tt, font=("Helvetica", 12))
				label.grid(row=7,column=0,padx=10,pady=5,columnspan=3)
				tab1_closed_observations.after(500, label.destroy)
				return

			title = proj_name + " - " + mod_name

			for i in observations_fr_project:
				y_count[i[2]] = y_count[i[2]] + 1

			y_count = { k:v for k, v in y_count.items() if v > 0 and k != 'RTM'}

			x = list(y_count.keys())
			y = list(y_count.values())

			ax.clear()
			now = datetime.datetime.now()
			date_display = now.strftime('%A, %d %B %Y, %H:%M')
			ax.set_title (title + "\n" + "Total number of findings per deliverable type\n(as of " + date_display + ")", fontsize=8)

			def func(pct, allvals):
				absolute = int(pct/100.*np.sum(allvals))
				return "{:.1f}%\n{:d}".format(pct, absolute)


			wedges, texts, autotexts = ax.pie(y, autopct=lambda pct: func(pct, y),
											  textprops=dict(color="w"))

			ax.legend(wedges, x,
					  title="Artifact Types",
					  loc="center left",
					  bbox_to_anchor=(1, 0, 0.5, 1))
			fig_artif.tight_layout()
			ax.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.

			#######################################
			''' GRAPH END '''
			#######################################

		def animate_finding_graph(i):
			#######################################
			''' GRAPH START '''
			#######################################

			## DATA
			x_cat = [x[1] for x in db.query_finding()]
			y_count = {}

			for i in x_cat:
				y_count[i] = 0

			artifact_name,proj_name,mod_name,team_name = artifact_type_select.get() ,proj_select.get(),proj_mod_select.get(),team_select.get()
			artif_name = artifact_name
			tim_name = team_name

			if(artifact_name == 'All'):
				artifact_name = db.query_artifact_name()
				artif_name = 'All Artifacts'
			else:
				artif_name = artifact_name
				artifact_name = [artifact_name]


			if(team_name == 'All'):
				team_name = db.query_team_name()
			else:
				team_name = [team_name]

			observations_fr_project = db.query_findings_by_projects(team_name,proj_name,mod_name,artifact_name,current_status)
			if(observations_fr_project == ''):
				ay.clear()
				tt = "No project modules in database for the selected project"
				label = Label(tab1_closed_observations, text=tt, font=("Helvetica", 12))
				label.grid(row=7,column=4,padx=10,pady=5,columnspan=3)
				tab1_closed_observations.after(500, label.destroy)
				return
			title = proj_name + " - " + mod_name

			for i in observations_fr_project:
				y_count[i[1]] = y_count[i[1]] + 1

			y_count = { k:v for k, v in y_count.items() if v > 0 and k != 'nan'}
			y_count = OrderedDict(sorted(y_count.items(), key=lambda kv: kv[1],reverse=True))

			if(len(list(y_count.keys())) >= 5):
				to_remove = list(y_count.keys())[5:]
				for x in to_remove:
					del y_count[x]

			x = list(y_count.values())
			y = list(y_count.keys())

			## SHOW
			ay.clear()
			bar_width = 0.4
			ay.barh(y,x,bar_width,color='yellow')
			ay.invert_yaxis()
			rects = ay.patches
			# print("rects",len(rects))
			labels = [ i for i in list(y_count.values())]
			# print("\n")

			for rect, label in zip(rects, labels):
				height = rect.get_height()/2
				width = rect.get_width() - 0.50
				ay.text(rect.get_x() + width, rect.get_y()+height,label,fontsize=8)

			ay.tick_params(
			axis='x',          # changes apply to the x-axis
			which='both',      # both major and minor ticks are affected
			bottom=False,      # ticks along the bottom edge are off
			top=False,         # ticks along the top edge are off
			labelbottom=False) # labels along the bottom edge are off

			ay.set_yticklabels(y,fontsize=6,wrap=True)

			for tick in ay.yaxis.get_major_ticks():
				tick.label1.set_verticalalignment('center')

			ay.set_title (title + "\n" + "Top 5 Deliverables (" + str(artif_name) +")", fontsize=8)

			#######################################
			## ''' GRAPH END '''
			#######################################

		def artifact_graph():
			# canvas_artifact = FigureCanvasTkAgg(fig_artif, master=tab1_closed_observations)
			# canvas_artifact.get_tk_widget().grid(row=7,column=0,padx=10,pady=5,columnspan=3)
			canvas_artifact.draw()  

			def file_save():
				animate_artifact_graph(0)
				canvas_artifact.draw() 
				orig_color = save_as_artifact_graph.cget("background")

				def change(orig_color):
					save_as_artifact_graph.configure(background = orig_color)

				filename =  filedialog.asksaveasfilename(initialdir = "/",title = "Select file",filetypes = (("PNG files","*.png"),("JPEG files","*.jpg"),("all files","*.*")))
				if os.path.basename(filename) is '':
					save_as_artifact_graph.configure(background = "red")
					tab1_closed_observations.after(1000, lambda: change(orig_color))
					return

				try:
					fig_artif.savefig(filename)
				except:
					save_as_artifact_graph.configure(background = "red")
					tab1_closed_observations.after(1000, lambda: change(orig_color))
					return

				save_as_artifact_graph.configure(background = "green")
				tt = "Exported"
				label = Label(tab1_closed_observations, text=tt, font=("Helvetica", 12))
				label.grid(row=0,column=0,padx=10,pady=5,columnspan=3)

				tab1_closed_observations.after(500, lambda: change(orig_color))
				tab1_closed_observations.after(500, label.destroy)
				print(filename)

			def export_win():
				output_files = current_output_folder + "/Closed/" + proj_select.get() + " -- " + proj_mod_select.get()
				
				animate_artifact_graph(0)
				canvas_artifact.draw() 

				orig_color = export_artifact_graph.cget("background")

				def change(orig_color):
					export_artifact_graph.configure(background = orig_color)

				if not os.path.exists(output_files):
					os.makedirs(output_files)

				try:
					fig_artif.savefig(output_files + "/artifact.png")
				except:
					export_artifact_graph.configure(background = "red")
					tab1_closed_observations.after(1000, lambda: change(orig_color))
					return

				export_artifact_graph.configure(background = "green")
				tt = "Exported"
				label = Label(tab1_closed_observations, text=tt, font=("Helvetica", 12))
				label.grid(row=0,column=0,padx=10,pady=5,columnspan=3)

				tab1_closed_observations.after(500, lambda: change(orig_color))
				tab1_closed_observations.after(500, label.destroy)

			export_artifact_graph = Button(tab1_closed_observations, text='Export', command=export_win)
			export_artifact_graph.grid(row=6,column=0,padx=30,pady=20,sticky='we')

			save_as_artifact_graph = Button(tab1_closed_observations, text='Save As', command=file_save)
			save_as_artifact_graph.grid(row=6,column=2,padx=5,pady=20,sticky='we')

		def finding_category():
			# canvas_finding = FigureCanvasTkAgg(fig_findin, master=tab1_closed_observations)
			# canvas_finding.get_tk_widget().grid(row=7,column=4,padx=10,pady=5,columnspan=3)
			canvas_finding.draw() 

			def file_save():
				animate_finding_graph(0)
				canvas_finding.draw() 
				orig_color = save_as_finding_graph.cget("background")

				def change(orig_color):
					save_as_finding_graph.configure(background = orig_color)

				filename =  filedialog.asksaveasfilename(initialdir = "/",title = "Select file",filetypes = (("PNG files","*.png"),("JPEG files","*.jpg"),("all files","*.*")))
				if os.path.basename(filename) is '':
					save_as_finding_graph.configure(background = "red")
					tab1_closed_observations.after(1000, lambda: change(orig_color))
					return

				try:
					fig_findin.savefig(filename)
				except:
					save_as_finding_graph.configure(background = "red")
					tab1_closed_observations.after(1000, lambda: change(orig_color))
					return

				save_as_finding_graph.configure(background = "green")
				tt = "Exported"
				label = Label(tab1_closed_observations, text=tt, font=("Helvetica", 12))
				label.grid(row=0,column=0,padx=10,pady=5,columnspan=3)
				tab1_closed_observations.after(500, lambda: change(orig_color))
				tab1_closed_observations.after(500, label.destroy)
				print(filename)

			def export_win():
				output_files = current_output_folder + "/Closed/" + proj_select.get() + " -- " + proj_mod_select.get()

				animate_finding_graph(0)
				canvas_finding.draw() 

				orig_color = export_finding_graph.cget("background")
				def change(orig_color):
					export_finding_graph.configure(background = orig_color)

				if not os.path.exists(output_files):
					os.makedirs(output_files)

				try:
					fig_findin.savefig(output_files + "/finding.png")
				except:
					export_finding_graph.configure(background = "red")
					tab1_closed_observations.after(1000, lambda: change(orig_color))
					return

				export_finding_graph.configure(background = "green")
				tt = "Exported"
				label = Label(tab1_closed_observations, text=tt, font=("Helvetica", 12))
				label.grid(row=0,column=0,padx=10,pady=5,columnspan=3)

				tab1_closed_observations.after(500, lambda: change(orig_color))
				tab1_closed_observations.after(500, label.destroy)

			export_finding_graph = Button(tab1_closed_observations, text='Export', command=export_win)
			export_finding_graph.grid(row=6,column=4,padx=30,pady=20,sticky='we')

			save_as_finding_graph = Button(tab1_closed_observations, text='Save As', command=file_save)
			save_as_finding_graph.grid(row=6,column=5,padx=5,pady=20,sticky='we')

		def update_graphs(a,b,c):
			animate_artifact_graph(0)
			animate_finding_graph(0)
			canvas_artifact.draw()  
			canvas_finding.draw() 
			# button_update_graphs.config(text="AUTO REFRESH is ON")

		def refresh_projects():
			projects = db.query_projects()
			if projects == []:
				print('No projects in database')

			proj_names = [project[1] for project in projects]
			proj_select.set(proj_names[0])
			project_dropdown = OptionMenu(tab1_closed_observations,proj_select,*proj_names)
			project_dropdown.configure(font='helvetica 12')
			project_dropdown.grid(row=1,column=1,padx=10,pady=10)

		def feedback_report():
			print(proj_location_select.get())
			lists = glob.glob(proj_location_select.get() + r'\*')
			if lists != []:
				latest_file = max(lists, key=os.path.getmtime)
			else:
				tt = 'There is no file under this folder'
				label = Label(tab1_closed_observations, text=tt, font=("Helvetica", 12), fg="red")
				label.grid(row=0,column=0,padx=10,pady=5)
				tab1_closed_observations.after(1000, label.destroy)
				return

			file_name = os.path.basename(latest_file)
			print(latest_file)
			copied_file_path = os.path.join(copied_folder, file_name)
			copyfile(latest_file, copied_file_path)
			os.system('start excel \"%s\"' % (copied_file_path))
			



		#Title
		closed_caption = Label(tab1_closed_observations, text='Closed/Deferred Observations')
		closed_caption.grid(row=0, column = 0,padx=10, pady=10)

		#Data Labels
		project_name_label = Label(tab1_closed_observations, text="Project Name")
		project_module_label = Label(tab1_closed_observations, text="Current Project Modules") 
		team_label = Label(tab1_closed_observations, text="Team Name")

		artifact_type_label = Label(tab1_closed_observations, text="Artifact Types") 
		finding_type_label = Label(tab1_closed_observations, text="Finding Category") 
		finding_label = Label(tab1_closed_observations, text="Findings")
		finding_detail_label = Label(tab1_closed_observations, text="Finding Details") 

		project_name_label.grid(row=1, column=0,padx=10, pady=10) 
		project_module_label.grid(row=2, column=0,padx=10, pady=10) 
		team_label.grid(row=3, column=0,padx=10, pady=10) 
		artifact_type_label.grid(row=4, column=0,padx=10, pady=10)
		finding_type_label.grid(row=5, column=0,padx=10, pady=10)
		finding_label.grid(row=0, column=3,padx=40, pady=10)
		finding_detail_label.grid(row=2, column=3,padx=40, pady=10)

		#String Vars Hidden Fields
		proj_select = StringVar(tab1_closed_observations,value='Default Text')
		proj_select.trace('w',update_modules)
		proj_select.trace('w',update_teams)
		proj_select.trace('w',update_artifacts)
		proj_select.trace('w',update_findings)
		proj_select.trace('w',update_listbox)
		proj_select.trace('w',update_graphs)

		proj_mod_select = StringVar(tab1_closed_observations,value='Default Text')
		proj_mod_select.trace('w',update_teams)
		proj_mod_select.trace('w',update_artifacts)
		proj_mod_select.trace('w',update_findings)
		proj_mod_select.trace('w',update_listbox)
		proj_mod_select.trace('w',update_graphs)

		team_select = StringVar(tab1_closed_observations,value='Default Text')
		team_select.trace('w',update_artifacts)
		team_select.trace('w',update_findings)
		team_select.trace('w',update_listbox)
		team_select.trace('w',update_graphs)

		artifact_type_select = StringVar(tab1_closed_observations,value='Default Text')
		artifact_type_select.trace('w',update_findings)
		artifact_type_select.trace('w',update_listbox)
		artifact_type_select.trace('w',update_graphs)

		finding_type_select = StringVar(tab1_closed_observations,value='Default Text')
		finding_type_select.trace('w',update_listbox)
		finding_type_select.trace('w',update_graphs)

		proj_location_select = StringVar(tab1_closed_observations,value='Default Text')

		#Dropdowns
		if(projects == []):
			print('No projects in database')
			return

		proj_names = [project[1] for project in projects]
		proj_select.set(proj_names[0])
		project_dropdown = OptionMenu(tab1_closed_observations,proj_select,*proj_names)
		project_dropdown.grid(row=1,column=1,padx=10,pady=10)

		artifact_graph()
		finding_category()
		
		button_update_graphs = Button(tab1_closed_observations, text="Refresh Projects", command=refresh_projects)
		button_update_graphs.grid(row=5, column=3,padx=40, pady=10)

		button_update_graphs = Button(tab1_closed_observations, text="Refresh Graphs", command=lambda:update_graphs(0,0,0))
		button_update_graphs.grid(row=6, column=3,padx=40, pady=10)

		button_open_feedback = Button(tab1_closed_observations, text="Feedback Report", command=feedback_report)
		button_open_feedback.grid(row=4, column=3,padx=40, pady=10)   
		button_open_feedback.bind()

class TabTwo(Frame):
	def __init__(self, parent):
		Frame.__init__(self, parent)
		self.tab2_note = ttk.Notebook(self,width=parent.winfo_screenwidth(), height=parent.winfo_screenheight())

		tab2_project_upload = ttk.Frame(self.tab2_note)
		tab2_project_module_upload = ttk.Frame(self.tab2_note)
		tab2_upload = ttk.Frame(self.tab2_note)

		self.tab2_note.add(tab2_project_upload, text= "Project Upload")
		self.tab2_note.add(tab2_project_module_upload, text= "Project Module Upload")
		self.tab2_note.add(tab2_upload, text= "Other Uploads")
		self.tab2_note.pack()

		self.tab_one_load(tab2_project_upload)
		self.tab_two_load(tab2_project_module_upload)
		self.tab_three_load(tab2_upload)
		widget_list = []

		widget_list.extend(tab2_project_upload.winfo_children())
		widget_list.extend(tab2_project_module_upload.winfo_children())
		widget_list.extend(tab2_upload.winfo_children())

		for wid in widget_list:
			try:
				wid.configure(font = 'helvetica 12')
			except:
				pass

	def tab_one_load(self,tab2_project_upload):
		######################################################################################
		#TAB2
		#TAB2 - TAB1
		######################################################################################
		projects = db.query_projects()

		def refresh_projects():
			projects = db.query_projects()
			if projects == []:
				print('No projects in database')

			proj_names = [project[1] for project in projects]
			proj_names.insert(0,"--")
			proj_select.set(proj_names[0])
			project_id_field = OptionMenu(tab2_project_upload,proj_select,*proj_names)
			project_id_field.configure(font='helvetica 12')
			project_id_field.grid(row=1, column=1,padx=10, pady=10)
			project_clarity_id_field.delete(0, END)
			project_name_field.delete(0, END)
			project_mgr_field.delete(0, END)
			project_qa_feedback_location_field.delete(0, END)
			project_clarity_id_field.focus_set()

		def update_name(a,b,c):
			if(proj_select.get() != '--'):
				proj_details = db.query_project_details(proj_select.get())
				project_clarity_id_var.set(str(proj_details[1]))
				project_name_var.set(str(proj_details[2]))
				project_mgr_var.set(str(proj_details[3]))
				project_qa_feedback_location_var.set(str(proj_details[4]))

				project_clarity_id_field.grid(row=2, column=1,padx=10, pady=10) 
				project_name_field.grid(row=3, column=1,padx=10, pady=10) 
				project_mgr_field.grid(row=4, column=1,padx=10, pady=10) 
				project_qa_feedback_location_field.grid(row=5, column=1,padx=10, pady=10)

		def project_upload():
			result = messagebox.askquestion("Create", "Are You Sure?", icon='warning')
			if result == 'yes':
				print ("Project about to be created")
				if (project_name_field.get() == "" or
					project_mgr_field.get() == "" or
					project_qa_feedback_location_field.get() == ""): 
					print("empty input") 
				else:
					if proj_select.get() == '--' :
						one_proj = [project_clarity_id_field.get(), project_name_field.get(), project_mgr_field.get(), str(project_qa_feedback_location_field.get()) + "\\"]
						two_proj = [project_clarity_id_var.get(), project_name_var.get(), project_mgr_var.get(), str(project_qa_feedback_location_var.get()) + "\\"]
						print(one_proj)
						print(two_proj)
						db.insert_project(one_proj)
						# set focus on the name_field box
						project_clarity_id_field.delete(0, END)
						project_name_field.delete(0, END)
						project_mgr_field.delete(0, END)
						project_qa_feedback_location_field.delete(0, END)
						project_clarity_id_field.focus_set()
						label = Label(tab2_project_upload, text="Successfully added!")
						label.grid(row=0, column=0, columnspan=2, rowspan=2, sticky=W+E+N+S, padx=5, pady=5)
						tab2_project_upload.after(2000, label.destroy)
					else:
						print("Can not update right now")

		lbl_project_title = Label(tab2_project_upload, text='New Project Upload?')
		lbl_project_title.grid(row=0, column = 0,padx=10, pady=10)

		project_id = Label(tab2_project_upload, text="Existing Project") 
		project_clarity_id = Label(tab2_project_upload, text="Clarity ID") 
		project_name = Label(tab2_project_upload, text="Project Name") 
		project_mgr = Label(tab2_project_upload, text="Project Manager") 
		project_qa_feedback_location = Label(tab2_project_upload, text="QA Feedback Location") 

		project_id.grid(row=1, column=0,padx=10, pady=10)
		project_clarity_id.grid(row=2, column=0,padx=10, pady=10) 
		project_name.grid(row=3, column=0,padx=10, pady=10) 
		project_mgr.grid(row=4, column=0,padx=10, pady=10) 
		project_qa_feedback_location.grid(row=5, column=0,padx=10, pady=10) 
		
		project_clarity_id_var = StringVar(tab2_project_upload,value='')
		project_clarity_id_var.trace('w',update_name)

		project_name_var = StringVar(tab2_project_upload,value='')
		project_name_var.trace('w',update_name)

		project_mgr_var = StringVar(tab2_project_upload,value='')
		project_mgr_var.trace('w',update_name)

		project_qa_feedback_location_var = StringVar(tab2_project_upload,value='')
		project_qa_feedback_location_var.trace('w',update_name)

		project_clarity_id_field = Entry(tab2_project_upload,textvariable=project_clarity_id_var)
		project_name_field = Entry(tab2_project_upload,width=50,textvariable=project_name_var)
		project_mgr_field = Entry(tab2_project_upload,width=50,textvariable=project_mgr_var)
		project_qa_feedback_location_field = Entry(tab2_project_upload,width=80,textvariable=project_qa_feedback_location_var)

		# project_clarity_id_field.grid(row=2, column=1,padx=10, pady=10) 
		# project_name_field.grid(row=3, column=1,padx=10, pady=10) 
		# project_mgr_field.grid(row=4, column=1,padx=10, pady=10) 
		# project_qa_feedback_location_field.grid(row=5, column=1,padx=10, pady=10) 

		proj_select = StringVar(tab2_project_upload)
		proj_select.trace('w',update_name)

		if projects == []:
			print('No projects in database')

		proj_names = [project[1] for project in projects]
		proj_names.insert(0,"--")
		proj_select.set(proj_names[0])
		project_id_field = OptionMenu(tab2_project_upload,proj_select,*proj_names)
		project_id_field.grid(row=1, column=1,padx=10, pady=10)

		project_upload_button = Button(tab2_project_upload, text="Create Project", command=project_upload)
		project_upload_button.grid(row =14, column =1,padx=10, pady=10)
		
		refresh_projects_button = Button(tab2_project_upload, text="Refresh Projects", command=refresh_projects)
		refresh_projects_button.grid(row=1, column=7,padx=10, pady=10, sticky=W+E+N+S, columnspan=2)

	def tab_two_load(self,tab2_project_module_upload):
		######################################################################################
		#TAB2
		#TAB2 - TAB2
		######################################################################################
		projects = db.query_projects()

		def refresh_projects():
			projects = db.query_projects()
			if projects == []:
				print('No projects in database')

			proj_names = [project[1] for project in projects]
			proj_select.set(proj_names[0])
			project_module_project_id_field = OptionMenu(tab2_project_module_upload,proj_select,*proj_names)
			project_module_current_modules_field = OptionMenu(tab2_project_module_upload,proj_mod_select,'')
			project_module_project_id_field.configure(font='helvetica 12')
			project_module_current_modules_field.configure(font='helvetica 12')
			project_module_project_id_field.grid(row=9, column=1,padx=10, pady=10)

		def update_options(a,b,c):
			# 2 # # Current project modules
			project_modules = db.query_project_modules(proj_select.get());
			if project_modules == []:
				print('No project modules in database for the selected project')
				label = Label(tab2_project_module_upload, text="No Modules in this project!")
				label.grid(row=10, column=4, columnspan=2, rowspan=2, sticky=W+E+N+S, padx=5, pady=5)
				tab2_project_module_upload.after(1500, label.destroy)

			# print(project_modules)
			proj_mo_names = [project[0] for project in project_modules]
			proj_mo_names.append('')
			proj_mod_select.set(proj_mo_names[0])
			project_module_current_modules_field = OptionMenu(tab2_project_module_upload,proj_mod_select,*proj_mo_names)
			project_module_current_modules_field.configure(font='helvetica 12')
			project_module_current_modules_field.grid(row=10, column=1,padx=10, pady=10)
			
		def update_name(a,b,c):
			project_module_name_field.grid(row=11, column=1,padx=10, pady=10) 

		def project_module_upload():
			global projects
			result = messagebox.askquestion("Create", "Are You Sure?", icon='warning')
			if result == 'yes':
				print("------")
				if (proj_select.get() == "" or
					project_module_name_field.get() == ""): 
					print("Empty input") 
				else:
					modules = []
					project_id = db.query_projects_id(proj_select.get())
					one_proj_module = [project_module_name_field.get(), project_id]
					modules.append(one_proj_module)
					db.insert_proj_module(modules)

					project_module_name_field.delete(0, END) 
					project_module_name_field.focus_set() 
					projects = db.query_projects()
					label = Label(tab2_project_module_upload, text="Successfully added!")
					label.grid(row=0, column=0, columnspan=2, rowspan=2, sticky=W+E+N+S, padx=5, pady=5)
					tab2_project_module_upload.after(2000, label.destroy)

		lbl_new_proj_module = Label(tab2_project_module_upload, text='New Project Module?')
		lbl_new_proj_module.grid(row=8, column = 0,padx=10, pady=10)

		project_module_project_id = Label(tab2_project_module_upload, text="Project_Name")
		project_module_current_modules = Label(tab2_project_module_upload, text="Current Project Modules") 
		project_module_name = Label(tab2_project_module_upload, text="Project Module Name") 

		project_module_project_id.grid(row=9, column=0,padx=10, pady=10) 
		project_module_current_modules.grid(row=10, column=0,padx=10, pady=10) 
		project_module_name.grid(row=11, column=0,padx=10, pady=10) 

		proj_select = StringVar(tab2_project_module_upload)
		proj_select.trace('w',update_options)
		proj_mod_select = StringVar(tab2_project_module_upload,value='Default Text')
		proj_mod_select.trace('w',update_name)

		project_module_name_field = Entry(tab2_project_module_upload,width=50,textvariable=proj_mod_select)

		if projects == []:
			print('No projects in database')

		proj_names = [project[1] for project in projects]
		proj_select.set(proj_names[0])
		project_module_project_id_field = OptionMenu(tab2_project_module_upload,proj_select,*proj_names)
		project_module_current_modules_field = OptionMenu(tab2_project_module_upload,proj_mod_select,'')
		project_module_project_id_field.grid(row=9, column=1,padx=10, pady=10)

		project_module_upload_button = Button(tab2_project_module_upload, text="Create Project Module", command=project_module_upload)
		project_module_upload_button.grid(row =14, column =1,padx=10, pady=10)

		refresh_projects_button = Button(tab2_project_module_upload, text="Refresh Projects", command=refresh_projects)
		refresh_projects_button.grid(row=1, column=5,padx=10, pady=10, sticky=W+E+N+S, columnspan=2, rowspan=3)


	def tab_three_load(self,tab2_upload):
		######################################################################################
		#TAB2
		#TAB2 - TAB3
		######################################################################################
		projects = db.query_projects()

		def refresh_projects():
			projects = db.query_projects()
			if projects == []:
				print('No projects in database')

			proj_names = [project[1] for project in projects]
			proj_select_upload.set(proj_names[0])
			drop = OptionMenu(tab2_upload,proj_select_upload,*proj_names)
			drop.configure(font='helvetica 12')
			drop.grid(row=9, column=2,padx=10, pady=10)

		def update_location(a,b,c):
			upload_project_location_field.grid(row=10, column=2,padx=10, pady=10) 

		def update_location_options(a,b,c):
			# print(a,b,c)
			print(proj_select_upload.get())
			query_projects_location = db.query_projects_location(proj_select_upload.get());
			proj_location_select.set(query_projects_location)

			g = ttk.Style()
			g.configure("green.Horizontal.TProgressbar", foreground='green', background='green')
			progress_bar = ttk.Progressbar(tab2_upload, style="green.Horizontal.TProgressbar", orient='horizontal',length=300, mode='determinate')
			progress_bar.grid(row=11, column=1,padx=10, pady=10, columnspan=2)
			progress_bar['maximum'] = 10
			pval = 0
			progress_bar['value'] = pval
			progress_bar.update_idletasks()

		def sync():
			g = ttk.Style()
			g.configure("green.Horizontal.TProgressbar", foreground='green', background='green')
			r = ttk.Style()
			r.configure("red.Horizontal.TProgressbar", foreground='red', background='red')
			progress_bar = ttk.Progressbar(tab2_upload, style="green.Horizontal.TProgressbar", orient='horizontal',length=300, mode='determinate')
			progress_bar.grid(row=11, column=1,padx=10, pady=10, columnspan=2)
			progress_bar['maximum'] = 10
			pval = 1
			progress_bar['value'] = pval
			progress_bar.update_idletasks()
			error = []
			for x in range(0,10):
				if(x == 0):
					# try:
					print ("Project selected:", proj_select_upload.get())
					project_name = proj_select_upload.get()
					Report([project_name])
					# except:
					# 	print("Error in uploading...")
					# 	error.append(project_name)
					# 	progress_bar.configure(style="red.Horizontal.TProgressbar")
					# 	continue
				pval = pval + 1 ## Or whatever increment makes sense.
				progress_bar['value'] = pval
				progress_bar.update_idletasks()

			print(error)

		def upload():
			query_projects_location = db.query_projects_location(proj_select_upload.get());
			file_name = filedialog.askopenfilename(parent=tab2_upload,initialdir=os.path.abspath(query_projects_location),
						   filetypes =(("All Files","*.*"),("Excel File", "*.xlsx"),("Excel File Old", "*.xls")),
						   title = "Choose a file."
						   )
			print (file_name)
			result = messagebox.askquestion("Upload External File", "Are you sure you want to continue? This could lead to potential problems in the records.", icon='warning')
			if result == 'yes':
				g = ttk.Style()
				g.configure("green.Horizontal.TProgressbar", foreground='green', background='green')
				r = ttk.Style()
				r.configure("red.Horizontal.TProgressbar", foreground='red', background='red')
				progress_bar = ttk.Progressbar(tab2_upload, style="green.Horizontal.TProgressbar", orient='horizontal',length=300, mode='determinate')
				progress_bar.grid(row=11, column=1,padx=10, pady=10, columnspan=2)
				progress_bar['maximum'] = 10
				pval = 1
				progress_bar['value'] = pval
				progress_bar.update_idletasks()
				error = []
				for x in range(0,10):
					if(x == 0):
						try:
							print ("Project selected:", proj_select_upload.get())
							project_name = proj_select_upload.get()
							Report([project_name],file_name)
						except:
							print("Error in uploading...")
							error.append(project_name)
							progress_bar.configure(style="red.Horizontal.TProgressbar")
							continue
					pval = pval + 1 ## Or whatever increment makes sense.
					progress_bar['value'] = pval
					progress_bar.update_idletasks()

				print(error)
				# try:
				# 	with open(file_name,'r') as UseFile:
				# 		tt = 'File Opening... %s' %(UseFile)
				# except:
				# 	print("Error")

		lbl_new_findings = Label(tab2_upload, text='Upload Latest File?')
		lbl_new_findings.grid(row=8, column = 0,padx=10, pady=10)
		
		upload_projects_dropdown = Label(tab2_upload, text="Project_Name")
		upload_project_location = Label(tab2_upload, text="Suggested Path") 
		upload_project_upload_location = Label(tab2_upload, text="Upload Status") 

		upload_projects_dropdown.grid(row=9, column=0,padx=10, pady=10) 
		upload_project_location.grid(row=10, column=0,padx=10, pady=10) 
		upload_project_upload_location.grid(row=11, column=0,padx=10, pady=10) 

		proj_select_upload = StringVar(tab2_upload)
		proj_select_upload.trace('w',update_location_options)
		proj_location_select = StringVar(tab2_upload,value='Default Text')
		proj_location_select.trace('w',update_location)

		upload_project_location_field = Entry(tab2_upload,width=90,textvariable=proj_location_select)
		upload_project_location_field.config(state=DISABLED)

		# # projects dropdown list
		if projects == []:
			print('No projects in database')
			label_error = Label(tab2_upload, text="No Projects in the database!")
			label_error.grid(row=0, column=0, columnspan=2, rowspan=2, sticky=W+E+N+S, padx=5, pady=5)
			tab2_upload.after(1500, label_error.destroy)

		proj_names = [project[1] for project in projects]
		proj_select_upload.set(proj_names[0])
		drop = OptionMenu(tab2_upload,proj_select_upload,*proj_names)
		drop.grid(row=9, column=2,padx=10, pady=10)

		button_suggested = Button(tab2_upload, text="Use Suggested File", command=sync)
		button_suggested.grid(row=10, column=4,padx=10, pady=10)

		button_file = Button(tab2_upload, text="Upload External File", command=upload)
		button_file.grid(row=11, column=4,padx=10, pady=10)

		refresh_projects_button = Button(tab2_upload, text="Refresh Projects", command=refresh_projects)
		refresh_projects_button.grid(row=1, column=5,padx=10, pady=10, sticky=W+E+N+S, columnspan=2, rowspan=3)

class TabThree(Frame):
	def __init__(self, parent):
		Frame.__init__(self, parent)
		self.tab3_note = ttk.Notebook(self,width=parent.winfo_screenwidth(), height=parent.winfo_screenheight())

		tab3_weekly_reports = ttk.Frame(self.tab3_note)
		tab3_yearly_reports = ttk.Frame(self.tab3_note)
		tab3_project_reports = ttk.Frame(self.tab3_note)

		self.tab3_note.add(tab3_weekly_reports, text= "Project Reports")
		self.tab3_note.add(tab3_yearly_reports, text= "Yearly Reports")
		self.tab3_note.add(tab3_project_reports, text= "Project Reports")

		self.tab3_note.pack()

		self.tab_one_load(tab3_weekly_reports)
		self.tab_two_load(tab3_yearly_reports)
		self.tab_three_load(tab3_project_reports)
		widget_list = []

		widget_list.extend(tab3_weekly_reports.winfo_children())
		widget_list.extend(tab3_yearly_reports.winfo_children())
		widget_list.extend(tab3_project_reports.winfo_children())

		for wid in widget_list:
			try:
				wid.configure(font = 'helvetica 12')
			except:
				pass

	def tab_one_load(self,tab3_weekly_reports):
		######################################################################################
		#TAB3
		#TAB3 - TAB1
		######################################################################################
		projects = db.query_projects();

		proj_select_charts_weekly = StringVar(tab3_weekly_reports)
		list_box_project = Listbox(tab3_weekly_reports,listvariable=proj_select_charts_weekly,selectmode='multiple',width=70,height = 10)

		def refresh_projects():
			projects = db.query_projects();
			if(projects == []):
			    print('No projects in database')
			    return

			proj_names = [project[1] for project in projects]
			proj_select_charts_weekly.set(proj_names)
			list_box_project.delete('0','end')
			list_box_project = Listbox(tab3_weekly_reports,listvariable=proj_select_charts_weekly,selectmode='multiple',width=70,height = 10)
			list_box_project.grid(row=11,column=2,padx=10, pady=10)

		def sync():
			selection = list_box_project.curselection()
			project_dict = {}
			for i in selection:
				proj = list_box_project.get(i)
				project_dict[proj] = db.query_projects_location(proj)
			report.set_vars(start_date_field_value.get(),end_date_field_value.get(),source_copy_field_value.get(),processed_file_path_value.get(),dir_path_value.get(),sheet_name_field_value.get(),QA_summary_file_name_value.get(),QA_review_file_name_value.get(),QA_review_file_sheet_value.get(),review_services_file_name_value.get())
			report.read(project_dict)
			label_suggested_path = Label(tab3_weekly_reports, text=QA_summary_file_name_value.get())
			label_suggested_path.grid(row=11, column=1, columnspan=2, rowspan=2)
			tab3_weekly_reports.after(2500, label_suggested_path.destroy)
			report.aakash_script_charts_pie()
			os.startfile(os.path.dirname(QA_summary_file_name_value.get()), 'open')
			os.startfile(os.path.join(os.path.dirname(QA_summary_file_name_value.get()),"Findings_Report_qa_review_summary.xlsx"), 'open')
			# file = filedialog.askopenfile(parent=tab3_weekly_reports,initialdir=,filetypes=[("Excel Files", "*.xlsx")],mode='rb',title='Check out file')


		chart_project_select_location = StringVar(tab3_weekly_reports,value='Default Text')

		lbl_new_findings = Label(tab3_weekly_reports, text='Weekly Project Reports')
		lbl_new_findings.grid(row=0, column = 0,padx=10, pady=10)
			
		start_date = Label(tab3_weekly_reports, text="Start Date") 
		end_date = Label(tab3_weekly_reports, text="End Date") 
		source_copy = Label(tab3_weekly_reports, text="Source Copy") 
		processed_file_path = Label(tab3_weekly_reports, text="Processed File Path")
		dir_path = Label(tab3_weekly_reports, text="Output Folder*")

		sheet_name = Label(tab3_weekly_reports, text="Sheet Name") 
		QA_summary_file_name = Label(tab3_weekly_reports, text="QA Summary File Name")
		QA_review_file_name = Label(tab3_weekly_reports, text="QA Review File Name")
		QA_review_file_sheet = Label(tab3_weekly_reports, text="QA Review File Sheet")
		review_services_file_name = Label(tab3_weekly_reports, text="Review Services File Name")
		upload_projects_dropdown = Label(tab3_weekly_reports, text="Project_Name")

		start_date.grid(row=1,column=0,padx=10,pady=10)
		end_date.grid(row=2,column=0,padx=10,pady=10)
		source_copy.grid(row=3,column=0,padx=10,pady=10)
		processed_file_path.grid(row=4,column=0,padx=10,pady=10)
		dir_path.grid(row=5,column=0,padx=10,pady=10)

		sheet_name.grid(row=6,column=0,padx=10,pady=10)
		QA_summary_file_name.grid(row=7,column=0,padx=10,pady=10)
		QA_review_file_name.grid(row=8,column=0,padx=10,pady=10)
		QA_review_file_sheet.grid(row=9,column=0,padx=10,pady=10)
		review_services_file_name.grid(row=10,column=0,padx=10,pady=10)
		upload_projects_dropdown.grid(row=11, column=0,padx=10, pady=10) 

		start_date_field_value = StringVar(tab3_weekly_reports,value='Default Text')
		start_date_field_value.set(default_date.strftime('%m/%d/%Y'))
		end_date_field_value = StringVar(tab3_weekly_reports,value='Default Text')
		end_date_field_value.set(now_date)

		source_copy_field_value = StringVar(tab3_weekly_reports,value='Default Text')
		source_copy_field_value.set('T:\\CITDR\\CITSQ\\QA\\Operational\\Status Reports\\QA reviews\\Access_DB\\CITSQ_Python_Tool\\EXCEL_REPORTS\\PROJECT_REPORTS\\SOURCE_COPY')
		# source_copy_field_value.set('C:\\Users\\ashah13\\OneDrive - WBG\\Desktop\\Python_Files\\Source_copy')

		processed_file_path_value = StringVar(tab3_weekly_reports,value='Default Text')
		processed_file_path_value.set('T:\\CITDR\\CITSQ\\QA\\Operational\\Status Reports\\QA reviews\\Access_DB\\CITSQ_Python_Tool\\EXCEL_REPORTS\\PROJECT_REPORTS\\PROCESSED')
		# processed_file_path_value.set('C:\\Users\\ashah13\\OneDrive - WBG\\Desktop\\Python_Files\\Processed')

		dir_path_value = StringVar(tab3_weekly_reports,value='Default Text')
		dir_path_value.set('T:\\CITDR\\CITSQ\\QA\\Operational\\Status Reports\\QA reviews\\Access_DB\\CITSQ_Python_Tool\\EXCEL_REPORTS\\PROJECT_REPORTS\\QA_REPORTS')
		# dir_path_value.set('C:\\Users\\ashah13\\OneDrive - WBG\\Desktop\\Python_Files\\new_QA Reports')

		sheet_name_field_value = StringVar(tab3_weekly_reports,value='Default Text')
		sheet_name_field_value.set('Test Plan,Test Cases,RTM,Test Summary Report, Test Results,Data Validations,Requirements')    
		# sheet_name_field_value.set('Test Plan,Test Cases,RTM,Test Summary Report, Test Results,Data Validations')    

		QA_summary_file_name_value = StringVar(tab3_weekly_reports,value='Default Text')
		QA_summary_file_name_value.set('T:\\CITDR\\CITSQ\\QA\\Operational\\Status Reports\\QA reviews\\Access_DB\\CITSQ_Python_Tool\\EXCEL_REPORTS\\PROJECT_REPORTS\\QA_REPORTS\\qa_review_summary.xlsx')
		# QA_summary_file_name_value.set('C:\\Users\\ashah13\\OneDrive - WBG\\Desktop\\Python_Files\\new_QA Reports\\qa_review_summary.xlsx')

		QA_review_file_name_value = StringVar(tab3_weekly_reports,value='Default Text')
		QA_review_file_name_value.set('T:\\CITDR\\CITSQ\\QA\\Operational\\Status Reports\\QA reviews\\Access_DB\\CITSQ_Python_Tool\\EXCEL_REPORTS\\PROJECT_REPORTS\\QA_REPORTS\\QA_Reviews.xlsx')
		# QA_review_file_name_value.set('C:\\Users\\ashah13\\OneDrive - WBG\\Desktop\\Python_Files\\new_QA Reports\\QA_Reviews.xlsx')

		QA_review_file_sheet_value = StringVar(tab3_weekly_reports,value='Default Text')
		QA_review_file_sheet_value.set('QA FY18 - Capital Proj')

		review_services_file_name_value = StringVar(tab3_weekly_reports,value='Default Text')
		review_services_file_name_value.set('T:\\CITDR\\CITSQ\\QA\\Operational\\Status Reports\\QA reviews\\Access_DB\\CITSQ_Python_Tool\\EXCEL_REPORTS\\PROJECT_REPORTS\\QA_REPORTS\\review_service.xlsx')
		# review_services_file_name_value.set('C:\\Users\\ashah13\\OneDrive - WBG\\Desktop\\Python_Files\\new_QA Reports\\review_service.xlsx')

		start_date_field = Entry(tab3_weekly_reports, width=50,textvariable=start_date_field_value)
		end_date_field = Entry(tab3_weekly_reports, width=50,textvariable=end_date_field_value)
		source_copy_field = Entry(tab3_weekly_reports, width=100,textvariable=source_copy_field_value)
		processed_file_path_field = Entry(tab3_weekly_reports, width=100,textvariable=processed_file_path_value)
		dir_path_field = Entry(tab3_weekly_reports, width=100,textvariable=dir_path_value)

		sheet_name_field = Entry(tab3_weekly_reports, width=70,textvariable=sheet_name_field_value)
		QA_summary_file_name_field = Entry(tab3_weekly_reports, width=70,textvariable=QA_summary_file_name_value)
		QA_review_file_name_field = Entry(tab3_weekly_reports, width=70,textvariable=QA_review_file_name_value)
		QA_review_file_sheet_field = Entry(tab3_weekly_reports, width=70,textvariable=QA_review_file_sheet_value)
		review_services_file_name_field = Entry(tab3_weekly_reports, width=70,textvariable=review_services_file_name_value)

		if projects == []:
			print('No projects in database')
			label_error = Label(tab3_weekly_reports, text="No Projects in the database!")
			label_error.grid(row=0, column=0, columnspan=2, rowspan=2, sticky=W+E+N+S, padx=5, pady=5)
			tab3_weekly_reports.after(1500, label_error.destroy)

		proj_names = [project[1] for project in projects]
		proj_select_charts_weekly.set(proj_names)

		# source_copy_field.config(state=DISABLED)
		# processed_file_path_field.config(state=DISABLED)

		sheet_name_field.config(state=DISABLED)
		QA_summary_file_name_field.config(state=DISABLED)
		QA_review_file_name_field.config(state=DISABLED)
		QA_review_file_sheet_field.config(state=DISABLED)
		review_services_file_name_field.config(state=DISABLED)

		start_date_field.grid(row=1,column=2,padx=10,pady=10)
		end_date_field.grid(row=2,column=2,padx=10,pady=10)
		source_copy_field.grid(row=3,column=2,padx=5,pady=5)
		processed_file_path_field.grid(row=4,column=2,padx=5,pady=5)
		dir_path_field.grid(row=5,column=2,padx=5,pady=5)

		sheet_name_field.grid(row=6,column=2,padx=5,pady=5)
		QA_summary_file_name_field.grid(row=7,column=2,padx=5,pady=5)
		QA_review_file_name_field.grid(row=8,column=2,padx=5,pady=5)
		QA_review_file_sheet_field.grid(row=9,column=2,padx=5,pady=5)
		review_services_file_name_field.grid(row=10,column=2,padx=5,pady=5)
		list_box_project.grid(row=11, column=2,padx=10, pady=10)

		button_update_graphs = Button(tab3_weekly_reports, text="Refresh Projects", command=refresh_projects)
		button_update_graphs.grid(row=9, column=4,padx=10, pady=10)

		button_suggested = Button(tab3_weekly_reports, text="Generate Reports", command=sync)
		button_suggested.grid(row=11, column=4,padx=10, pady=10)

	def tab_two_load(self,tab3_yearly_reports):
		######################################################################################
		#TAB3
		#TAB3 - TAB3
		######################################################################################
		def sync_yearly():
			print("here")
			report.set_quarter_vars(source_copy_field_value.get(),dir_path_value.get(),int(number_sheets_value.get()),int(top_n_value.get()))
			# label_suggested_path = Label(tab3_yearly_reports, text=QA_summary_file_name_value.get())
			# label_suggested_path.grid(row=11, column=1, columnspan=2, rowspan=2)
			# tab3_yearly_reports.after(2500, label_suggested_path.destroy)
			report.aakash_script_charts_column()
			os.startfile(os.path.dirname(dir_path_value.get()), 'open')
			os.startfile(os.path.dirname(dir_path_value.get()), 'open')
			os.startfile(os.path.join(os.path.dirname(dir_path_value.get()),"bar_report_qa-quarter-yearly.xlsx"), 'open')
			# file = filedialog.askopenfile(parent=tab3_yearly_reports,initialdir=,filetypes=[("Excel Files", "*.xlsx")],mode='rb',title='Check out file')

		def button_copy():
			r = Tk()
			r.withdraw()
			r.clipboard_clear()
			r.clipboard_append(dir_path_value.get())
			r.update() # now it stays on the clipboard after the window is closed
			r.destroy()

		chart_project_select_location = StringVar(tab3_yearly_reports,value='Default Text')

		lbl_new_findings = Label(tab3_yearly_reports, text='Yearly Project Reports')
		lbl_new_findings.grid(row=0, column = 0,padx=10, pady=10)
			

		source_copy = Label(tab3_yearly_reports, text="Input File") 
		dir_path = Label(tab3_yearly_reports, text="Output Folder*")
		number_sheets = Label(tab3_yearly_reports, text="Number of Sheets (FY)")
		top_n = Label(tab3_yearly_reports, text="Top N findings")

		source_copy.grid(row=3,column=0,padx=10,pady=10)
		dir_path.grid(row=5,column=0,padx=10,pady=10)
		number_sheets.grid(row=6,column=0,padx=10,pady=10)
		top_n.grid(row=7,column=0,padx=10,pady=10)

		source_copy_field_value = StringVar(tab3_yearly_reports,value='Default Text')
		source_copy_field_value.set('T:\\CITDR\\CITSQ\\QA\\Operational\\Status Reports\\QA reviews\\Access_DB\\CITSQ_Python_Tool\\EXCEL_REPORTS\\YEARLY_REPORTS\\qa-quarter-yearly.xlsx')

		dir_path_value = StringVar(tab3_yearly_reports,value='Default Text')
		dir_path_value.set('T:\\CITDR\\CITSQ\\QA\\Operational\\Status Reports\\QA reviews\\Access_DB\\CITSQ_Python_Tool\\EXCEL_REPORTS\\YEARLY_REPORTS\\')

		number_sheets_value = StringVar(tab3_yearly_reports,value='Default Text')
		number_sheets_value.set('3')

		top_n_value = StringVar(tab3_yearly_reports,value='Default Text')
		top_n_value.set('2')

		source_copy_field = Entry(tab3_yearly_reports, width=100,textvariable=source_copy_field_value)
		dir_path_field = Entry(tab3_yearly_reports, width=100,textvariable=dir_path_value)
		number_sheets_field = Entry(tab3_yearly_reports, width=10,textvariable=number_sheets_value)
		top_n_field = Entry(tab3_yearly_reports, width=10,textvariable=top_n_value)

		projects = db.query_projects();
		if projects == []:
			print('No projects in database')
			label_error = Label(tab3_yearly_reports, text="No Projects in the database!")
			label_error.grid(row=0, column=0, columnspan=2, rowspan=2, sticky=W+E+N+S, padx=5, pady=5)
			tab3_yearly_reports.after(1500, label_error.destroy)

		source_copy_field.grid(row=3,column=2,padx=5,pady=5)
		dir_path_field.grid(row=5,column=2,padx=5,pady=5)
		number_sheets_field.grid(row=6,column=2,padx=5,pady=5)
		top_n_field.grid(row=7,column=2,padx=5,pady=5)

		button_suggested = Button(tab3_yearly_reports, text="Generate Reports", command=sync_yearly)
		button_suggested.grid(row=11, column=4,padx=10, pady=10)

		button_cc = Button(tab3_yearly_reports, text="[][]", command=button_copy)
		button_cc.grid(row=5, column=3,padx=2, pady=2)

	def tab_three_load(self,tab3_project_reports):
		######################################################################################
		#TAB2
		#TAB2 - TAB2
		######################################################################################
		projects = db.query_projects();
		current_status = [1,2,3,4]
		fig_artif = Figure(figsize=(5,3))
		fig_findin = Figure(figsize=(6,3))

		ax = fig_artif.add_subplot(111)
		ay = fig_findin.add_subplot(111)

		listbox_observation_select = StringVar(tab3_project_reports,value='Default Text')
		list_box_closed = Listbox(tab3_project_reports,listvariable=listbox_observation_select,selectmode='browse',width=60,height=4)
		
		canvas_artifact = FigureCanvasTkAgg(fig_artif, master=tab3_project_reports)
		canvas_artifact.get_tk_widget().grid(row=7,column=0,padx=10,pady=5,columnspan=3)

		canvas_finding = FigureCanvasTkAgg(fig_findin, master=tab3_project_reports)
		canvas_finding.get_tk_widget().grid(row=7,column=4,padx=10,pady=5,columnspan=3)

		def display_record(record):
			a,b,c,d,e,f,g,h,i,j,k,l,m,n,o = record

			#Header
			a_doc_name_and_ver = a
			b_test_count = b
			c_project_name = c
			d_module_name = d
			f_finding_type = f
			o_artifact_type = o
			n_finding_id = n

			#Review
			e_location_in_artifact = e
			g_QA_review = g
			i_review_date = i

			#Response
			h_project_response = h
			follow_up_reviews = [j,k,l,m]

			follow_up_text = ''
			for idx,date_text in enumerate(follow_up_reviews):
				if(date_text <= default_date):
					follow_up_reviews[idx] = ' '
				else:
					follow_up_text = follow_up_text + 'Follow Up %d:\t%s\n' % ((idx+1),follow_up_reviews[idx].strftime('%b %d %Y %I:%M%p'))

			header_text = "Finding ID:\n%s\n\nDoc Name & Version:\n%s\n\nTest Count:\n%s\n\nProject Name:\n%s\n\nModule Name:\n%s\n\nFinding Type:\n%s\n\nArtifact Name:\n%s\n\n" % (n,a,b,c,d,f,o)
			review_text = "\nArtifact Location:\n%s\n\nQA Review:\n%s\n\nReview Date:\n%s\n" % (e,g,i.strftime('%b %d %Y %I:%M %p'))
			response_text = "\nProject Response:\n%s\n\n" % (h)
			response_text = response_text + follow_up_text

			full_text = header_text + review_text + response_text
			width, height = 60,1
			text_response = scrolledtext.ScrolledText(tab3_project_reports,width=width, height=height, wrap='word')
			text_response.insert(1.0, full_text)
			text_response['font'] = ('consolas', '10','bold')
			text_response.grid(row=2,column=4,padx=20,pady=10,rowspan=4,sticky="news")

		def list_has_changed(self):
			### CHANGE THIS IF SELECTMODE IS MULTIPLE
			values = [list_box_closed.get(idx) for idx in list_box_closed.curselection()]
			if values != []:
				item = values[0]
				a,b,c,d = item.split("  ->  ")
				b = int(b)
				query_result = db.query_findings_finding_id(b)
				if(query_result != ''):
					display_record(query_result[0])

		def update_modules(a,b,c):
			# 2 # # Current project modules
			proj_name = proj_select.get()

			query_projects_location = db.query_projects_location(proj_name);
			proj_location_select.set(query_projects_location)

			project_modules = db.query_project_modules(proj_name);
			if project_modules == []:
				print('No project modules in database for the selected project')
				label = Label(tab3_project_reports, text="No Modules in this project!")
				label.grid(row=0, column=0, sticky=W+E+N+S, padx=5, pady=5)
				tab3_project_reports.after(500, label.destroy)
				return

			# print(project_modules)
			proj_mo_names = [module[0] for module in project_modules]
			proj_mod_select.set(proj_mo_names[0])
			project_module_dropdown = OptionMenu(tab3_project_reports,proj_mod_select,*proj_mo_names)
			project_module_dropdown.configure(font='helvetica 12')
			project_module_dropdown.grid(row=2, column=1,padx=10, pady=10)

		def update_teams(a,b,c):
			teams = db.query_teams()
			if(teams == []):
				print('No Teams found')
				return

			team_list = [team[0] for team in teams]
			team_list.insert(0,'All')
			team_select.set(team_list[0])
			team_dropdown = OptionMenu(tab3_project_reports,team_select,*team_list)
			team_dropdown.configure(font='helvetica 12')
			team_dropdown.grid(row=3,column=1,padx=10,pady=10)

		def update_artifacts(a,b,c):
			team_name,proj_name,mod_name = team_select.get(),proj_select.get(),proj_mod_select.get()

			if(team_name == 'All'):
				team_name = db.query_team_name()
			else:
				team_name = [team_name]

			artifacts_fr_project = db.query_artifacts_details(team_name,proj_name,mod_name,current_status)
			title_text = proj_name + " - " + mod_name
			
			if artifacts_fr_project == []:
				print('No artifacts in database')
				label_error = Label(tab3_project_reports, text="No artifacts in the database!")
				label_error.grid(row=0, column=0, columnspan=2, rowspan=2, sticky=W+E+N+S, padx=5, pady=5)
				tab3_project_reports.after(1500, label_error.destroy)
				return 

			artifact_names = [artifact[0] for artifact in artifacts_fr_project]
			if artifact_names != []:
				artifact_names.insert(0,'All')
				artifact_type_select.set(artifact_names[0])
				artifact_type_dropdown = OptionMenu(tab3_project_reports,artifact_type_select,*artifact_names)
				artifact_type_dropdown.configure(font='helvetica 12')
				artifact_type_dropdown.grid(row=4, column=1,padx=10, pady=10)

		def update_findings(a,b,c):
			artifact_name,team_name,proj_name,mod_name = artifact_type_select.get(),team_select.get(),proj_select.get(),proj_mod_select.get()
			if(artifact_name == 'All'):
				artifact_name = db.query_artifact_name()
			else:
				artifact_name = [artifact_name]

			if(team_name == 'All'):
				team_name = db.query_team_name()
			else:
				team_name = [team_name]

			observations_fr_project = db.query_findings(team_name,proj_name,mod_name,artifact_name,current_status)
			
			if observations_fr_project == []:
				print('No findings in database')
				label_error = Label(tab3_project_reports, text="No findings in the database!")
				label_error.grid(row=0, column=0, columnspan=2, rowspan=2, sticky=W+E+N+S, padx=5, pady=5)
				tab3_project_reports.after(500, label_error.destroy)
				return 

			# max_len = 40
			finding_names = [finding[0] for finding in observations_fr_project]
			# finding_names = [finding[0][:max_len] + '..' if len(finding[0]) > max_len else finding[0] for finding in observations_fr_project ]

			if finding_names != []:
				finding_names.insert(0,'All')
				finding_type_select.set(finding_names[0])
				finding_type_dropdown = OptionMenu(tab3_project_reports,finding_type_select,*finding_names)
				finding_type_dropdown.configure(font='helvetica 12')
				finding_type_dropdown.grid(row=5, column=1,padx=10, pady=10)
				tab3_project_reports.grid_columnconfigure(1, weight=1)
				tab3_project_reports.grid_columnconfigure(4, weight=1)

		def update_listbox(a,b,c):
			finding_name,artifact_name,team_name,proj_name,mod_name = finding_type_select.get(),artifact_type_select.get(),team_select.get(),proj_select.get(),proj_mod_select.get()
			if(artifact_name == 'All'):
				artifact_name = db.query_artifact_name()
			else:
				artifact_name = [artifact_name]

			if(finding_name == 'All'):
				finding_name = db.query_finding_name()
			else:
				finding_name = [finding_name]

			if(team_name == 'All'):
				team_name = db.query_team_name()
			else:
				team_name = [team_name]

			listbox_for_findings = db.query_findings_details(team_name,proj_name,mod_name,artifact_name,finding_name,current_status)
			if listbox_for_findings == []:
				print('No projects in database')
				label_error = Label(tab3_project_reports, text="No Findings in the database!")
				label_error.grid(row=0, column=0, columnspan=2, rowspan=2, sticky=W+E+N+S, padx=5, pady=5)
				tab3_project_reports.after(500, label_error.destroy)
				return

			findings_list = [str(find[1]) + "  ->  " + str(find[0]) + "  ->  " + str(find[4]) + "  ->  " + str(find[2]) for find in listbox_for_findings]
			listbox_observation_select.set(findings_list)
			list_box_closed.grid(row=0, column=4,padx=20,pady=10, rowspan=2)
			list_box_closed.bind("<Double-Button-1>", list_has_changed)
			
		def animate_artifact_graph(i):
			#######################################
			''' GRAPH START '''
			#######################################
			## DATA
			x_cat = [x[1] for x in db.query_artifact()]
			x_cat = x_cat[:-2]

			y_count = {}
			for i in x_cat:
				y_count[i] = 0

			artifact_name,proj_name,mod_name,team_name = artifact_type_select.get() ,proj_select.get(),proj_mod_select.get(),team_select.get()
			if(artifact_name == 'All'):
				artifact_name = db.query_artifact_name()
			else:
				artifact_name = [artifact_name]

			if(team_name == 'All'):
				team_name = db.query_team_name()
			else:
				team_name = [team_name]

			observations_fr_project = db.query_findings_by_projects(team_name,proj_name,mod_name,artifact_name,current_status)
			if(observations_fr_project == ''):
				ax.clear()
				tt = "No project modules in database for the selected project"
				label = Label(tab3_project_reports, text=tt, font=("Helvetica", 12))
				label.grid(row=7,column=0,padx=10,pady=5,columnspan=3)
				tab3_project_reports.after(500, label.destroy)
				return

			title = proj_name + " - " + mod_name

			for i in observations_fr_project:
				y_count[i[2]] = y_count[i[2]] + 1

			y_count = { k:v for k, v in y_count.items() if v > 0 and k != 'RTM'}

			x = list(y_count.keys())
			y = list(y_count.values())

			ax.clear()
			now = datetime.datetime.now()
			date_display = now.strftime('%A, %d %B %Y, %H:%M')
			ax.set_title (title + "\n" + "Total number of findings per deliverable type\n(as of " + date_display + ")", fontsize=8)

			def func(pct, allvals):
				absolute = int(pct/100.*np.sum(allvals))
				return "{:.1f}%\n{:d}".format(pct, absolute)


			wedges, texts, autotexts = ax.pie(y, autopct=lambda pct: func(pct, y),
											  textprops=dict(color="w"))

			ax.legend(wedges, x,
					  title="Artifact Types",
					  loc="center left",
					  bbox_to_anchor=(1, 0, 0.5, 1))
			fig_artif.tight_layout()
			ax.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.

			#######################################
			''' GRAPH END '''
			#######################################

		def animate_finding_graph(i):
			#######################################
			''' GRAPH START '''
			#######################################

			## DATA
			x_cat = [x[1] for x in db.query_finding()]
			y_count = {}

			for i in x_cat:
				y_count[i] = 0

			artifact_name,proj_name,mod_name,team_name = artifact_type_select.get() ,proj_select.get(),proj_mod_select.get(),team_select.get()
			artif_name = artifact_name
			tim_name = team_name

			if(artifact_name == 'All'):
				artifact_name = db.query_artifact_name()
				artif_name = 'All Artifacts'
			else:
				artif_name = artifact_name
				artifact_name = [artifact_name]


			if(team_name == 'All'):
				team_name = db.query_team_name()
				# print(team_name)
			else:
				team_name = [team_name]

			observations_fr_project = db.query_findings_by_projects(team_name,proj_name,mod_name,artifact_name,current_status)
			if(observations_fr_project == ''):
				ay.clear()
				tt = "No project modules in database for the selected project"
				label = Label(tab3_project_reports, text=tt, font=("Helvetica", 12))
				label.grid(row=7,column=4,padx=10,pady=5,columnspan=3)
				tab3_project_reports.after(500, label.destroy)
				return
			title = proj_name + " - " + mod_name

			for i in observations_fr_project:
				y_count[i[1]] = y_count[i[1]] + 1

			y_count = { k:v for k, v in y_count.items() if v > 0 and k != 'nan'}
			y_count = OrderedDict(sorted(y_count.items(), key=lambda kv: kv[1],reverse=True))

			if(len(list(y_count.keys())) >= 5):
				to_remove = list(y_count.keys())[5:]
				for x in to_remove:
					del y_count[x]

			x = list(y_count.values())
			y = list(y_count.keys())

			## SHOW
			ay.clear()  
			bar_width = 0.4
			ay.barh(y,x,bar_width,color='yellow')
			ay.invert_yaxis()
			rects = ay.patches
			# print("rects",len(rects))
			labels = [ i for i in list(y_count.values())]
			# print("\n")

			for rect, label in zip(rects, labels):
				height = rect.get_height()/2
				width = rect.get_width() - 0.50
				ay.text(rect.get_x() + width, rect.get_y()+height,label,fontsize=8)

			ay.tick_params(
			axis='x',          # changes apply to the x-axis
			which='both',      # both major and minor ticks are affected
			bottom=False,      # ticks along the bottom edge are off
			top=False,         # ticks along the top edge are off
			labelbottom=False) # labels along the bottom edge are off

			ay.set_yticklabels(y,fontsize=6,wrap=True)

			for tick in ay.yaxis.get_major_ticks():
				tick.label1.set_verticalalignment('center')

			ay.set_title (title + "\n" + "Top 5 Deliverables (" + str(artif_name) +")", fontsize=8)

			#######################################
			## ''' GRAPH END '''
			#######################################

		def artifact_graph():
			# canvas_artifact = FigureCanvasTkAgg(fig_artif, master=tab3_project_reports)
			# canvas_artifact.get_tk_widget().grid(row=7,column=0,padx=10,pady=5,columnspan=3)
			canvas_artifact.draw()  

			def file_save():
				animate_artifact_graph(0)
				canvas_artifact.draw() 
				orig_color = save_as_artifact_graph.cget("background")

				def change(orig_color):
					save_as_artifact_graph.configure(background = orig_color)

				filename =  filedialog.asksaveasfilename(initialdir = "/",title = "Select file",filetypes = (("PNG files","*.png"),("JPEG files","*.jpg"),("all files","*.*")))
				if os.path.basename(filename) is '':
					save_as_artifact_graph.configure(background = "red")
					tab3_project_reports.after(1000, lambda: change(orig_color))
					return
				try:
					fig_artif.savefig(filename)
				except:
					save_as_artifact_graph.configure(background = "red")
					tab3_project_reports.after(1000, lambda: change(orig_color))
					return

				save_as_artifact_graph.configure(background = "green")
				tt = "Exported"
				label = Label(tab3_project_reports, text=tt, font=("Helvetica", 12))
				label.grid(row=0,column=0,padx=10,pady=5,columnspan=3)

				tab3_project_reports.after(500, lambda: change(orig_color))
				tab3_project_reports.after(500, label.destroy)
				print(filename)

			def export_win():
				output_files = current_output_folder + "/All/" + proj_select.get() + " -- " + proj_mod_select.get()

				animate_artifact_graph(0)
				canvas_artifact.draw() 

				orig_color = export_artifact_graph.cget("background")

				def change(orig_color):
					export_artifact_graph.configure(background = orig_color)

				if not os.path.exists(output_files):
					os.makedirs(output_files)
				try:
					fig_artif.savefig(output_files + "/artifact.png")
				except:
					export_artifact_graph.configure(background = "red")
					tab3_project_reports.after(1000, lambda: change(orig_color))
					return

				export_artifact_graph.configure(background = "green")
				tt = "Exported"
				label = Label(tab3_project_reports, text=tt, font=("Helvetica", 12))
				label.grid(row=0,column=0,padx=10,pady=5,columnspan=3)

				tab3_project_reports.after(500, lambda: change(orig_color))
				tab3_project_reports.after(500, label.destroy)

			export_artifact_graph = Button(tab3_project_reports, text='Export', command=export_win)
			export_artifact_graph.grid(row=6,column=0,padx=30,pady=20,sticky='we')

			save_as_artifact_graph = Button(tab3_project_reports, text='Save As', command=file_save)
			save_as_artifact_graph.grid(row=6,column=2,padx=5,pady=20,sticky='we')

		def finding_category():
			# canvas_finding = FigureCanvasTkAgg(fig_findin, master=tab3_project_reports)
			# canvas_finding.get_tk_widget().grid(row=7,column=4,padx=10,pady=5,columnspan=3)
			canvas_finding.draw() 

			def file_save():
				animate_finding_graph(0)
				canvas_finding.draw() 
				orig_color = save_as_finding_graph.cget("background")

				def change(orig_color):
					save_as_finding_graph.configure(background = orig_color)

				filename =  filedialog.asksaveasfilename(initialdir = "/",title = "Select file",filetypes = (("PNG files","*.png"),("JPEG files","*.jpg"),("all files","*.*")))
				if os.path.basename(filename) is '':
					save_as_finding_graph.configure(background = "red")
					tab3_project_reports.after(1000, lambda: change(orig_color))
					return

				try:
					fig_findin.savefig(filename)
				except:
					save_as_finding_graph.configure(background = "red")
					tab3_project_reports.after(1000, lambda: change(orig_color))
					return

				save_as_finding_graph.configure(background = "green")
				tt = "Exported"
				label = Label(tab3_project_reports, text=tt, font=("Helvetica", 12))
				label.grid(row=0,column=0,padx=10,pady=5,columnspan=3)
				tab3_project_reports.after(500, lambda: change(orig_color))
				tab3_project_reports.after(500, label.destroy)
				print(filename)

			def export_win():
				output_files = current_output_folder + "/All/" + proj_select.get() + " -- " + proj_mod_select.get()

				animate_finding_graph(0)
				canvas_finding.draw() 

				orig_color = export_finding_graph.cget("background")
				def change(orig_color):
					export_finding_graph.configure(background = orig_color)

				if not os.path.exists(output_files):
					os.makedirs(output_files)

				try:
					fig_findin.savefig(output_files + "/finding.png")
				except:
					export_finding_graph.configure(background = "red")
					tab3_project_reports.after(1000, lambda: change(orig_color))
					return

				export_finding_graph.configure(background = "green")
				tt = "Exported"
				label = Label(tab3_project_reports, text=tt, font=("Helvetica", 12))
				label.grid(row=0,column=0,padx=10,pady=5,columnspan=3)

				tab3_project_reports.after(500, lambda: change(orig_color))
				tab3_project_reports.after(500, label.destroy)

			export_finding_graph = Button(tab3_project_reports, text='Export', command=export_win)
			export_finding_graph.grid(row=6,column=4,padx=30,pady=20,sticky='we')

			save_as_finding_graph = Button(tab3_project_reports, text='Save As', command=file_save)
			save_as_finding_graph.grid(row=6,column=5,padx=5,pady=20,sticky='we')

		def update_graphs(a,b,c):
			animate_artifact_graph(0)
			animate_finding_graph(0)
			canvas_artifact.draw()  
			canvas_finding.draw() 
			# button_update_graphs.config(text="AUTO REFRESH is ON")

		def refresh_projects():
			projects = db.query_projects();
			if(projects == []):
			    print('No projects in database')
			    return

			proj_names = [project[1] for project in projects]
			proj_select.set(proj_names[0])
			project_dropdown = OptionMenu(tab3_project_reports,proj_select,*proj_names)
			project_dropdown.configure(font='helvetica 12')
			project_dropdown.grid(row=1,column=1,padx=10,pady=10)

		def feedback_report():
			lists = glob.glob(proj_location_select.get() + r'\*')
			if lists != []:
				latest_file = max(lists, key=os.path.getmtime)
			else:
				tt = 'There is no file under this folder'
				label = Label(tab3_project_reports, text=tt, font=("Helvetica", 12), fg="red")
				label.grid(row=0,column=0,padx=10,pady=5)
				tab3_project_reports.after(1000, label.destroy)
				return

			file_name = os.path.basename(latest_file)
			print(latest_file)
			copied_file_path = os.path.join(copied_folder, file_name)
			copyfile(latest_file, copied_file_path)
			os.system('start excel \"%s\"' % (copied_file_path))
			
		#Title
		closed_caption = Label(tab3_project_reports, text='Closed/Deferred Observations')
		closed_caption.grid(row=0, column = 0,padx=10, pady=10)

		#Data Labels
		project_name_label = Label(tab3_project_reports, text="Project Name")
		project_module_label = Label(tab3_project_reports, text="Current Project Modules") 
		team_label = Label(tab3_project_reports, text="Team Name")

		artifact_type_label = Label(tab3_project_reports, text="Artifact Types") 
		finding_type_label = Label(tab3_project_reports, text="Finding Category") 
		finding_label = Label(tab3_project_reports, text="Findings")
		finding_detail_label = Label(tab3_project_reports, text="Finding Details") 

		project_name_label.grid(row=1, column=0,padx=10, pady=10) 
		project_module_label.grid(row=2, column=0,padx=10, pady=10) 
		team_label.grid(row=3, column=0,padx=10, pady=10) 
		artifact_type_label.grid(row=4, column=0,padx=10, pady=10)
		finding_type_label.grid(row=5, column=0,padx=10, pady=10)
		finding_label.grid(row=0, column=3,padx=40, pady=10)
		finding_detail_label.grid(row=2, column=3,padx=40, pady=10)

		#String Vars Hidden Fields
		proj_select = StringVar(tab3_project_reports,value='Default Text')
		proj_select.trace('w',update_modules)
		proj_select.trace('w',update_teams)
		proj_select.trace('w',update_artifacts)
		proj_select.trace('w',update_findings)
		proj_select.trace('w',update_listbox)
		proj_select.trace('w',update_graphs)

		proj_mod_select = StringVar(tab3_project_reports,value='Default Text')
		proj_mod_select.trace('w',update_teams)
		proj_mod_select.trace('w',update_artifacts)
		proj_mod_select.trace('w',update_findings)
		proj_mod_select.trace('w',update_listbox)
		proj_mod_select.trace('w',update_graphs)

		team_select = StringVar(tab3_project_reports,value='All')
		team_select.trace('w',update_artifacts)
		team_select.trace('w',update_findings)
		team_select.trace('w',update_listbox)
		team_select.trace('w',update_graphs)

		artifact_type_select = StringVar(tab3_project_reports,value='All')
		artifact_type_select.trace('w',update_findings)
		artifact_type_select.trace('w',update_listbox)
		artifact_type_select.trace('w',update_graphs)

		finding_type_select = StringVar(tab3_project_reports,value='Default Text')
		finding_type_select.trace('w',update_listbox)
		finding_type_select.trace('w',update_graphs)

		proj_location_select = StringVar(tab3_project_reports,value='Default Text')

		#Dropdowns
		if(projects == []):
			print('No projects in database')
			return

		proj_names = [project[1] for project in projects]
		proj_select.set(proj_names[0])
		project_dropdown = OptionMenu(tab3_project_reports,proj_select,*proj_names)
		project_dropdown.grid(row=1,column=1,padx=10,pady=10)

		artifact_graph()
		finding_category()


		button_update_graphs = Button(tab3_project_reports, text="Refresh Projects", command=refresh_projects)
		button_update_graphs.grid(row=5, column=3,padx=40, pady=10)

		button_update_graphs = Button(tab3_project_reports, text="Refresh Graphs", command=lambda:update_graphs(0,0,0))
		button_update_graphs.grid(row=6, column=3,padx=40, pady=10)

		button_open_feedback = Button(tab3_project_reports, text="Feedback Report", command=feedback_report)
		button_open_feedback.grid(row=4, column=3,padx=20, pady=20)   
		button_open_feedback.bind()

class TabFour(Frame):
	def __init__(self, parent):
		Frame.__init__(self, parent)
		self.tab4_note = ttk.Notebook(self,width=parent.winfo_screenwidth(), height=parent.winfo_screenheight())

		tab4_status = ttk.Frame(self.tab4_note)
		tab4_severity = ttk.Frame(self.tab4_note)
		tab4_artifact = ttk.Frame(self.tab4_note)
		tab4_finding = ttk.Frame(self.tab4_note)
		tab5_db = ttk.Frame(self.tab4_note)

		self.tab4_note.add(tab4_status, text= "Resolution Status")
		self.tab4_note.add(tab4_severity, text= "Severity")
		self.tab4_note.add(tab4_artifact, text= "Artifact Type")
		self.tab4_note.add(tab4_finding, text= "Finding Category")
		self.tab4_note.add(tab5_db, text= "DB Path")

		self.tab4_note.pack()

		self.tab_one_load(tab4_status)
		self.tab_two_load(tab4_severity)
		self.tab_three_load(tab4_artifact)
		self.tab_four_load(tab4_finding)
		self.tab_five_load(tab5_db)
		widget_list = []

		widget_list.extend(tab4_status.winfo_children())
		widget_list.extend(tab4_severity.winfo_children())
		widget_list.extend(tab4_artifact.winfo_children())
		widget_list.extend(tab4_finding.winfo_children())
		widget_list.extend(tab5_db.winfo_children())

		for wid in widget_list:
			try:
				wid.configure(font = 'helvetica 12')
			except:
				pass

	def tab_one_load(self,tab4_status):
		######################################################################################
		#TAB4
		#TAB4 - TAB1
		######################################################################################
		status_results = db.query_resolution_status();
		tree_status = ttk.Treeview(tab4_status)

		def status_add():        
			result = messagebox.askquestion("Create Status", "Are You Sure?", icon='warning')
			if result == 'yes':
				print ("Status about to be created")
				status = status_name_field.get()
				db.insert_status(status)

				top = db.query_status_top()
				tree_status.insert("" , 0, values=(top[0][0],top[0][1]))
				tree_status.grid(row=7, column=2,sticky=W+E+N+S, padx=20, pady=30)

		lbl_resolution_status = Label(tab4_status, text='Resolution Status')
		lbl_resolution_status.grid(row=0, column = 0,padx=10, pady=10)
			
		status_name = Label(tab4_status, text="Status Name") 
		status_name.grid(row=1,column=0,padx=10,pady=10)
		status_name_field_value = StringVar(tab4_status,value='Default Text')
		status_name_field_value.set('Urgent')
		status_name_field = Entry(tab4_status, width=50,textvariable=status_name_field_value)
		status_name_field.grid(row=1,column=2,padx=10,pady=10)
		
		button_create_status = Button(tab4_status, text="Create New Item", command=status_add)
		button_create_status.grid(row=1, column=4,padx=10, pady=10)

		tree_status["columns"]=("id","name")
		tree_status['show'] = 'headings'
		tree_status.column("id", anchor='center',width=100 )
		tree_status.column("name", anchor='center',width=100)
		tree_status.heading("id", text="Status ID")
		tree_status.heading("name", text="Status Name")
		for x in status_results:
			tree_status.insert("" , 0, values=(x[0],x[1]))
		tree_status.grid(row=7, column=2, sticky=W+E+N+S, padx=20, pady=30)


	def tab_two_load(self,tab4_severity):
		######################################################################################
		#TAB4
		#TAB4 - TAB2
		######################################################################################
		severity_results = db.query_severity();
		tree_severity = ttk.Treeview(tab4_severity)

		def severity_add():        
			result = messagebox.askquestion("Create Severity Item", "Are You Sure?", icon='warning')
			if result == 'yes':
				print ("Severity about to be created")
				severity = severity_name_field.get()
				db.insert_severity(severity)

				top = db.query_severity_top()
				tree_severity.insert("" , 0, values=(top[0][0],top[0][1]))
				tree_severity.grid(row=7, column=2,sticky=W+E+N+S, padx=20, pady=30)

		lbl_resolution_severity = Label(tab4_severity, text='Severity')
		lbl_resolution_severity.grid(row=0, column = 0,padx=10, pady=10)
			
		severity_name = Label(tab4_severity, text="Severity Name") 
		severity_name.grid(row=1,column=0,padx=10,pady=10)
		severity_name_field_value = StringVar(tab4_severity,value='Default Text')
		severity_name_field_value.set('Urgent')
		severity_name_field = Entry(tab4_severity, width=50,textvariable=severity_name_field_value)
		severity_name_field.grid(row=1,column=2,padx=10,pady=10)
		
		button_create_severity = Button(tab4_severity, text="Create New Item", command=severity_add)
		button_create_severity.grid(row=1, column=4,padx=10, pady=10)

		tree_severity["columns"]=("id","name")
		tree_severity['show'] = 'headings'
		tree_severity.column("id", anchor='center',width=100 )
		tree_severity.column("name", anchor='center',width=100)
		tree_severity.heading("id", text="Severity ID")
		tree_severity.heading("name", text="Severity Name")
		for x in severity_results:
			tree_severity.insert("" , 0, values=(x[0],x[1]))
		tree_severity.grid(row=7, column=2, sticky=W+E+N+S, padx=20, pady=30)


	def tab_three_load(self,tab4_artifact):
		print("Tab Three")
		######################################################################################
		#TAB4
		#TAB4 - TAB3
		######################################################################################
		artifact_results = db.query_artifact();
		tree_artifact = ttk.Treeview(tab4_artifact)

		def artifact_add():        
			result = messagebox.askquestion("Create Artifact Item", "Are You Sure?", icon='warning')
			if result == 'yes':
				print ("Artifact about to be created")
				artifact = artifact_name_field.get()
				db.insert_artifact(artifact)

				top = db.query_artifact_top()
				tree_artifact.insert("" , 0, values=(top[0][0],top[0][1]))
				tree_artifact.grid(row=7, column=2,sticky=W+E+N+S, padx=20, pady=30)

		lbl_resolution_artifact = Label(tab4_artifact, text='Artifact Types')
		lbl_resolution_artifact.grid(row=0, column = 0,padx=10, pady=10)
			
		artifact_name = Label(tab4_artifact, text="Artifact Type Name") 
		artifact_name.grid(row=1,column=0,padx=10,pady=10)
		artifact_name_field_value = StringVar(tab4_artifact,value='Default Text')
		artifact_name_field_value.set('Test Evidences')
		artifact_name_field = Entry(tab4_artifact, width=50,textvariable=artifact_name_field_value)
		artifact_name_field.grid(row=1,column=2,padx=10,pady=10)
		
		button_create_artifact = Button(tab4_artifact, text="Create New Item", command=artifact_add)
		button_create_artifact.grid(row=1, column=4,padx=10, pady=10)

		tree_artifact["columns"]=("id","name")
		tree_artifact['show'] = 'headings'
		tree_artifact.column("id", anchor='center',width=100 )
		tree_artifact.column("name", anchor='center',width=100)
		tree_artifact.heading("id", text="Artifact Type ID")
		tree_artifact.heading("name", text="Artifact Name")
		for x in artifact_results:
			tree_artifact.insert("" , 0, values=(x[0],x[1]))
		tree_artifact.grid(row=7, column=2, sticky=W+E+N+S, padx=20, pady=30)


	def tab_four_load(self,tab4_finding):
		print("Tab Four")
		######################################################################################
		#TAB4
		#TAB4 - TAB4
		######################################################################################
		finding_results = db.query_finding();
		tree_finding = ttk.Treeview(tab4_finding)

		def finding_add():        
			result = messagebox.askquestion("Create Finding Category", "Are You Sure?", icon='warning')
			if result == 'yes':
				print ("Finding Category about to be created")
				finding = finding_name_field.get()
				artifact_nam = finding_artifact_name_field_value.get()
				db.insert_finding_category(finding,artifact_nam)

				top = db.query_finding_top()
				tree_finding.insert("" , 0, values=(top[0][0],top[0][1],top[0][2]))
				tree_finding.grid(row=7, column=2,sticky=W+E+N+S, padx=20, pady=30)

		lbl_resolution_finding = Label(tab4_finding, text='Finding Category')
		lbl_resolution_finding.grid(row=0, column = 0,padx=10, pady=10)
			
		finding_name = Label(tab4_finding, text="Finding Category Name") 
		finding_name.grid(row=1,column=0,padx=10,pady=10)

		finding_artifact_name = Label(tab4_finding, text="Artifact Type") 
		finding_artifact_name.grid(row=2,column=0,padx=10,pady=10)

		finding_name_field_value = StringVar(tab4_finding,value='Default Text')
		finding_name_field_value.set('Test Evidences')

		finding_name_field = Entry(tab4_finding, width=50,textvariable=finding_name_field_value)
		finding_name_field.grid(row=1,column=2,padx=10,pady=10)
			
		finding_artifact_name_field_value = StringVar(tab4_finding,value='Default Text')
		artifact_results = db.query_artifact()
		artifact_names = [artifact[1] for artifact in artifact_results]
		finding_artifact_name_field_value.set(artifact_names[0])
		finding_artifact_name_field = OptionMenu(tab4_finding,finding_artifact_name_field_value,*artifact_names)    
		finding_artifact_name_field.grid(row=2,column=2,padx=10,pady=10)

		button_create_finding = Button(tab4_finding, text="Create New Item", command=finding_add)
		button_create_finding.grid(row=1, column=4,padx=10, pady=10)

		tree_finding["columns"]=("id","name","a_id")
		tree_finding['show'] = 'headings'
		tree_finding.column("id", anchor='center',width=150 )
		tree_finding.column("name", anchor='center',width=150)
		tree_finding.column("a_id", anchor='center',width=150)

		tree_finding.heading("id", text="Finding Category ID")
		tree_finding.heading("name", text="Finding Category Name")
		tree_finding.heading("a_id", text="Artifact ID")

		for x in finding_results:
			tree_finding.insert("" , 0, values=(x[0],x[1],x[2]))
		
		tree_finding.grid(row=7, column=2, sticky=W+E+N+S, padx=20, pady=30)

	def tab_five_load(self,tab5_database):
		######################################################################################
		#TAB4
		#TAB4 - TAB5
		######################################################################################
		def db_update_path():
			result = messagebox.askquestion("Create", "Are You Sure?", icon='warning')
			if result == 'yes':
				default_dir = default_dir_value.get()
				dbq = mdb_file_path_value.get()
				db.setup_conn(default_dir,dbq)
				tt = "DB Connected!"
				label = Label(tab5_database, text=tt, font=("Helvetica", 12),fg="green")
				label.grid(row=0,column=0,padx=10,pady=5)
				tab5_database.after(1500, label.destroy)

		def button_copy():
			r = Tk()
			r.withdraw()
			r.clipboard_clear()
			r.clipboard_append(mdb_file_path_value.get())
			r.update() # now it stays on the clipboard after the window is closed
			r.destroy()

		change_path_title = Label(tab5_database, text='Change DB Path?')
		change_path_title.grid(row=8, column = 0,padx=10, pady=10)

		default_dir_label = Label(tab5_database, text="Defualt Directory") 
		mdb_file_path_label = Label(tab5_database, text="MDB File Path") 
		
		default_dir_label.grid(row=9, column=0,padx=10, pady=10) 
		mdb_file_path_label.grid(row=10, column=0,padx=10, pady=10) 
	   
		default_dir_value = StringVar(tab5_database,value='Default Text')
		default_dir_value.set(default_dir)
		mdb_file_path_value = StringVar(tab5_database,value='Default Text')
		mdb_file_path_value.set(dbq)

		default_dir_field = Entry(tab5_database,width=100,textvariable=default_dir_value)
		mdb_file_path_field = Entry(tab5_database,width=100,textvariable=mdb_file_path_value)
		default_dir_field.grid(row=9, column=1,padx=10, pady=10) 
		mdb_file_path_field.grid(row=10, column=1,padx=10, pady=10) 
		

		button_cc = Button(tab5_database, text="[][]", command=button_copy)
		button_cc.grid(row=10, column=2,padx=2, pady=2)

		db_path_button = Button(tab5_database, text="Update DB Path", command=db_update_path)
		db_path_button.grid(row =14, column =1,padx=10, pady=10)

class MainWindow(Frame):
	def __init__(self, window, **kwargs):
		Frame.__init__(self, window, **kwargs)

		self.pad=3
		self.window = window
		self._geom='200x200+0+0'
		window.geometry("{0}x{1}+0+0".format(window.winfo_screenwidth()-self.pad, window.winfo_screenheight()-self.pad))
		window.bind('<Escape>',self.toggle_geom)
		self.load_ui()

	def load_ui(self):
		self.note = ttk.Notebook(self,width=self.window.winfo_screenwidth()-(2*self.pad), height=self.window.winfo_screenheight()-(2*self.pad))

		self.tab1 = TabOne(self.note)
		self.tab2 = TabTwo(self.note)
		self.tab3 = TabThree(self.note)
		self.tab4 = TabFour(self.note)

		self.note.add(self.tab1, text = "Current Status")
		self.note.add(self.tab2, text = "Uploads")
		self.note.add(self.tab3, text = "View Charts")
		self.note.add(self.tab4, text = "Settings")

		self.note.pack()

	def toggle_geom(self,event):
		geom=self.window.winfo_geometry()
		print(geom,self._geom)
		self.window.geometry(self._geom)
		self._geom=geom

class Splash(Toplevel):
	def __init__(self, window,**kwargs):
		Toplevel.__init__(self, window,**kwargs)
		self.overrideredirect(True)
		self.pad = 3
		self.lift()
		x = self.winfo_screenwidth()-self.pad
		y = self.winfo_screenheight()-self.pad
		self.geometry("{0}x{1}+0+0".format(x,y))

		self.canvas = Canvas(self, width = x, height =y,bg='white')    
		self.canvas.pack(expand=YES, fill=BOTH,anchor=CENTER)

		# self.frames = [PhotoImage(file=image_loader_name,format = 'gif -index %i' %(i)) for i in range(4)]
		for x in range(5):
			self.img = PhotoImage(file=image_loader_name,format = 'gif -index %d' %(x))
			width = self.canvas.winfo_width()
			height = self.canvas.winfo_height()
			if(width != 1 and height != 1):
				self.canvas.create_image(width/2, height/2, anchor=CENTER, image=self.img, tags="bg_img")
				time.sleep(1)
			self.update()
		## required to make window show before the program gets to the mainloop
		self.update()


'''

MainWindow : End

'''

######################## Main ###################################

def main():
	window = Tk()
	window.withdraw()
	splash = Splash(window)

	# Set window property
	mygreen = "#d2ffd2"
	myred = "#dd0202"
	window.configure(background=mygreen) 
	window.title('CITSQ Productivity Tool')
	# Set style property
	window.style = ttk.Style()
	window.style.theme_create( "MyStyle", parent="default", 
			settings={
				"TNotebook": {
						"configure": { "tabmargins": [10, 10, 0, 0] } 
				 },
				 # Padding for making tabs bigger
				"TNotebook.Tab": {
						"configure": {"padding": [50, 10] },
						 "map":       {
							"background": [("selected", '#F0F8FF')],
							"font": [("selected",'helvetica 13'),("disabled",'helvetica 13')],
							"expand": [("selected", [1, 1, 1, 0])] 
						  } 

				}
			}
	)
	window.style.theme_use("MyStyle")
	# ('winnative', 'clam', 'alt', 'default', 'classic', 'vista', 'xpnative')
	app = MainWindow(window)
	## simulate a delay while loading
	## finished loading so destroy splash
	splash.destroy()
	## show window again
	window.deiconify()
	app.pack(side="top", fill="both", expand=True)

	window.mainloop()

if __name__ == '__main__':
	main()
